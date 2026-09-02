#!/usr/bin/env python3
"""
EC2 Audit Script
- Finds over-allocated EC2 instances (high vCPU/RAM vs actual usage)
- Finds long-idle/unused EC2 instances via CloudWatch metrics
- Detects last human access via NetworkPacketsIn spikes + CloudTrail SSH/SSM events
- Shows which ports are exposed (Security Groups) and which were actually hit (VPC Flow Logs)
- Reports owner/contact tags so you know who to reach out to
"""

import boto3
import json
import time
from datetime import datetime, timezone, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Configuration ──────────────────────────────────────────────────────────────
LOOKBACK_DAYS = 30          # CloudWatch look-back window
IDLE_CPU_THRESHOLD = 5.0    # % avg CPU – below this = idle
OVERALLOC_CPU_THRESHOLD = 20.0  # % avg CPU – below this = over-allocated
IDLE_NETWORK_THRESHOLD = 10_000  # bytes/day avg below which = idle
IDLE_DAYS_THRESHOLD = 14    # instance with NO data for this many days = long idle

# NetworkPacketsIn: daily packet count below this is "background noise only"
# (ARP, health checks, keepalives). Real SSH/HTTP sessions produce far more.
BACKGROUND_PACKETS_PER_DAY = 2_000

# CloudTrail lookback for SSH/SSM events (max 90 days for free tier)
CLOUDTRAIL_LOOKBACK_DAYS = 90

# Tags to look for owner/contact info (checked in order)
OWNER_TAG_KEYS = ["Owner", "owner", "Contact", "contact",
                  "Email", "email", "CreatedBy", "created-by",
                  "Team", "team", "User", "user"]

SKIP_REGIONS: set[str] = set()

# ── Helpers ────────────────────────────────────────────────────────────────────

def get_tag(tags: list, key: str, default: str = "N/A") -> str:
    if not tags:
        return default
    for t in tags:
        if t.get("Key", "").lower() == key.lower():
            return t.get("Value", default)
    return default

def get_owner(tags: list) -> str:
    for key in OWNER_TAG_KEYS:
        val = get_tag(tags, key)
        if val != "N/A":
            return f"{key}={val}"
    return "N/A"

def format_age(dt: datetime) -> str:
    if dt is None:
        return "unknown"
    delta = datetime.now(timezone.utc) - dt
    days = delta.days
    if days < 1:
        return f"{delta.seconds // 3600}h ago"
    if days < 365:
        return f"{days}d ago"
    return f"{days // 365}y {days % 365}d ago"

def get_metric_avg(cw_client, instance_id: str, metric_name: str,
                   namespace: str = "AWS/EC2",
                   stat: str = "Average",
                   period_days: int = LOOKBACK_DAYS) -> float | None:
    """Return the average of a CloudWatch metric over the lookback window, or None."""
    end = datetime.now(timezone.utc)
    start = end - timedelta(days=period_days)
    try:
        resp = cw_client.get_metric_statistics(
            Namespace=namespace,
            MetricName=metric_name,
            Dimensions=[{"Name": "InstanceId", "Value": instance_id}],
            StartTime=start,
            EndTime=end,
            Period=3600 * 24,
            Statistics=[stat],
        )
        datapoints = resp.get("Datapoints", [])
        if not datapoints:
            return None
        return sum(d[stat] for d in datapoints) / len(datapoints)
    except Exception:
        return None

def get_last_datapoint_time(cw_client, instance_id: str,
                             metric_name: str = "CPUUtilization",
                             period_days: int = 455) -> datetime | None:
    """Return the most-recent timestamp that had a CloudWatch data point (daily granularity)."""
    end = datetime.now(timezone.utc)
    start = end - timedelta(days=period_days)
    try:
        resp = cw_client.get_metric_statistics(
            Namespace="AWS/EC2",
            MetricName=metric_name,
            Dimensions=[{"Name": "InstanceId", "Value": instance_id}],
            StartTime=start,
            EndTime=end,
            Period=86400,
            Statistics=["Average"],
        )
        datapoints = resp.get("Datapoints", [])
        if not datapoints:
            return None
        return max(d["Timestamp"] for d in datapoints)
    except Exception:
        return None

def get_last_significant_network_day(cw_client, instance_id: str,
                                      period_days: int = 455) -> tuple[datetime | None, int]:
    """
    Return (most-recent day with real traffic, packet count on that day).
    Filters out ARP/health-check noise (<= BACKGROUND_PACKETS_PER_DAY packets/day).
    Returns (None, 0) when no meaningful traffic found.
    """
    end = datetime.now(timezone.utc)
    start = end - timedelta(days=period_days)
    try:
        resp = cw_client.get_metric_statistics(
            Namespace="AWS/EC2",
            MetricName="NetworkPacketsIn",
            Dimensions=[{"Name": "InstanceId", "Value": instance_id}],
            StartTime=start,
            EndTime=end,
            Period=86400,
            Statistics=["Sum"],
        )
        datapoints = resp.get("Datapoints", [])
        active = [(d["Timestamp"], int(d["Sum"])) for d in datapoints
                  if d["Sum"] > BACKGROUND_PACKETS_PER_DAY]
        if not active:
            return None, 0
        best = max(active, key=lambda x: x[0])
        return best[0], best[1]
    except Exception:
        return None, 0

# ── CloudTrail: SSH / SSM access per region ───────────────────────────────────

# Event names that indicate a human logged into or connected to an instance
_SSH_SSM_EVENTS = {
    "SendSSHPublicKey",   # EC2 Instance Connect (ssh via browser/CLI)
    "StartSession",       # SSM Session Manager – starts a shell session
    "ResumeSession",      # SSM Session Manager – resumes a session
    "StartPortForwardingSession",          # SSM port-forward
    "StartPortForwardingSessionToRemoteHost",
}

def get_cloudtrail_last_access(region: str) -> dict[str, dict]:
    """
    Scan CloudTrail in `region` for SSH/SSM login events over the last
    CLOUDTRAIL_LOOKBACK_DAYS days.

    Returns:
        { instance_id: {
            "time":       datetime,
            "event":      str,   # e.g. "SendSSHPublicKey"
            "user":       str,   # IAM user/role that triggered it
            "source_ip":  str,   # caller's IP address
          } }
    Only keeps the most-recent event per instance.
    """
    ct = boto3.client("cloudtrail", region_name=region)
    end   = datetime.now(timezone.utc)
    start = end - timedelta(days=CLOUDTRAIL_LOOKBACK_DAYS)
    last_access: dict[str, dict] = {}

    for event_name in _SSH_SSM_EVENTS:
        try:
            paginator = ct.get_paginator("lookup_events")
            pages = paginator.paginate(
                LookupAttributes=[{"AttributeKey": "EventName", "AttributeValue": event_name}],
                StartTime=start,
                EndTime=end,
            )
            for page in pages:
                for event in page.get("Events", []):
                    event_time = event.get("EventTime")
                    if not event_time:
                        continue

                    # Parse the full CloudTrail JSON for richer details
                    user = "unknown"
                    source_ip = "unknown"
                    raw = event.get("CloudTrailEvent", "")
                    found_iids: set[str] = set()

                    if raw:
                        try:
                            payload = json.loads(raw)
                            source_ip = payload.get("sourceIPAddress", "unknown")
                            # IAM identity
                            identity = payload.get("userIdentity", {})
                            user = (identity.get("userName")
                                    or identity.get("sessionContext", {})
                                       .get("sessionIssuer", {}).get("userName")
                                    or identity.get("arn", "unknown").split("/")[-1])
                            # Instance ID from requestParameters
                            req = payload.get("requestParameters") or {}
                            iid = req.get("instanceId") or req.get("target", "").split("/")[0]
                            if iid and iid.startswith("i-"):
                                found_iids.add(iid)
                        except Exception:
                            pass

                    # Also scan Resources list
                    for res in event.get("Resources", []):
                        if res.get("ResourceType") == "AWS::EC2::Instance":
                            found_iids.add(res["ResourceName"])

                    for iid in found_iids:
                        existing = last_access.get(iid)
                        if existing is None or event_time > existing["time"]:
                            last_access[iid] = {
                                "time":      event_time,
                                "event":     event_name,
                                "user":      user,
                                "source_ip": source_ip,
                            }
        except Exception:
            pass  # CloudTrail might not be enabled or accessible

    return last_access

# ── Port visibility helpers ────────────────────────────────────────────────────

def get_sg_ports(ec2_client, sg_ids: list[str]) -> str:
    """
    Return a compact string of inbound port rules from the given security groups.
    Only includes rules that allow traffic from the internet (0.0.0.0/0 or ::/0).
    Example: "tcp/22, tcp/80, tcp/443, udp/53"
    """
    if not sg_ids:
        return "—"
    try:
        resp = ec2_client.describe_security_groups(GroupIds=sg_ids)
    except Exception:
        return "—"

    public_cidrs = {"0.0.0.0/0", "::/0"}
    ports: list[str] = []
    seen: set[str] = set()

    for sg in resp.get("SecurityGroups", []):
        for perm in sg.get("IpPermissions", []):
            proto = perm.get("IpProtocol", "")
            # Only include rules with at least one public-facing source
            sources = (
                [r["CidrIp"]  for r in perm.get("IpRanges", [])]
                + [r["CidrIpv6"] for r in perm.get("Ipv6Ranges", [])]
            )
            if not any(s in public_cidrs for s in sources):
                # Also check for rules that have NO cidr restriction
                # (prefixlists/security-group sources – treat as internal, skip)
                continue

            if proto == "-1":
                tag = "all-traffic"
            elif proto in ("tcp", "udp"):
                fp = perm.get("FromPort")
                tp = perm.get("ToPort")
                if fp is None:
                    continue
                tag = f"{proto}/{fp}" if fp == tp else f"{proto}/{fp}-{tp}"
            elif proto == "icmp":
                tag = "icmp"
            else:
                tag = proto

            if tag not in seen:
                seen.add(tag)
                ports.append(tag)

    return ", ".join(sorted(ports)) if ports else "no public rules"


def get_vpc_flow_log_groups(ec2_client, vpc_ids: set) -> dict:
    """
    Return {vpc_id: log_group_name} for VPCs that have VPC Flow Logs
    delivered to CloudWatch Logs.  Only considers ACTIVE flow logs.
    """
    if not vpc_ids:
        return {}
    try:
        resp = ec2_client.describe_flow_logs(
            Filters=[
                {"Name": "resource-id",          "Values": list(vpc_ids)},
                {"Name": "log-destination-type",  "Values": ["cloud-watch-logs"]},
                {"Name": "flow-log-status",       "Values": ["ACTIVE"]},
            ]
        )
    except Exception:
        return {}

    vpc_to_group: dict[str, str] = {}
    for fl in resp.get("FlowLogs", []):
        rid = fl.get("ResourceId", "")
        lg  = fl.get("LogGroupName", "")
        if rid and lg and rid not in vpc_to_group:
            vpc_to_group[rid] = lg
    return vpc_to_group


def query_flow_log_ports_bulk(
    logs_client,
    log_group: str,
    eni_ids: set,
    lookback_days: int = LOOKBACK_DAYS,
    timeout_sec: int = 45,
) -> dict:
    """
    Run ONE CloudWatch Logs Insights query against `log_group` to find, for
    every ENI in `eni_ids`, which destination ports received ACCEPT traffic.

    Returns: { eni_id: "22 (1.2k), 443 (890), 8080 (44)" }

    Uses the default VPC Flow Log v2 format (14 space-separated fields).
    Silently returns {} on any error or timeout.
    """
    if not eni_ids:
        return {}

    end   = datetime.now(timezone.utc)
    start = end - timedelta(days=lookback_days)

    # Parse the 14-field default v2 flow log line; field positions:
    #  0=ver 1=acct 2=iface 3=src 4=dst 5=sport 6=dport 7=proto
    #  8=pkts 9=bytes 10=start 11=end 12=action 13=status
    query = """fields @message
| parse @message "* * * * * * * * * * * * * *" as _v,_a,iface,_s,_d,_sp,dport,_pr,_pk,_by,_t1,_t2,action,_st
| filter action = "ACCEPT"
| stats count(*) as hits by iface, dport
| sort hits desc
| limit 1000"""

    try:
        start_resp = logs_client.start_query(
            logGroupName=log_group,
            startTime=int(start.timestamp()),
            endTime=int(end.timestamp()),
            queryString=query,
        )
        query_id = start_resp["queryId"]
    except Exception:
        return {}

    # Poll for results
    deadline = time.time() + timeout_sec
    while time.time() < deadline:
        time.sleep(2)
        try:
            result = logs_client.get_query_results(queryId=query_id)
        except Exception:
            return {}
        status = result.get("status", "")
        if status in ("Complete", "Failed", "Cancelled"):
            break
    else:
        # Timed out – cancel silently
        try:
            logs_client.stop_query(queryId=query_id)
        except Exception:
            pass
        return {}

    if result.get("status") != "Complete":
        return {}

    # Build {eni_id: {port: hit_count}}
    from collections import defaultdict
    eni_ports: dict = defaultdict(dict)
    for row in result.get("results", []):
        fields = {f["field"]: f["value"] for f in row}
        iface = fields.get("iface", "")
        dport = fields.get("dport", "")
        hits  = int(fields.get("hits", 0))
        if iface in eni_ids and dport.isdigit():
            eni_ports[iface][dport] = hits

    # Format each ENI's top ports
    output: dict[str, str] = {}
    for eni, port_map in eni_ports.items():
        top = sorted(port_map.items(), key=lambda x: -x[1])[:8]
        parts = []
        for port, cnt in top:
            if cnt >= 1_000_000:
                cnt_str = f"{cnt/1_000_000:.1f}M"
            elif cnt >= 1_000:
                cnt_str = f"{cnt/1_000:.1f}k"
            else:
                cnt_str = str(cnt)
            parts.append(f"{port}({cnt_str})")
        output[eni] = ", ".join(parts)
    return output


# ── Per-region worker ──────────────────────────────────────────────────────────

def audit_region(region: str) -> dict:
    """Audit all running EC2 instances in one region. Returns a result dict."""
    result = {
        "region": region,
        "instances": [],
        "error": None,
    }
    try:
        session = boto3.session.Session(region_name=region)
        ec2   = session.client("ec2")
        cw    = session.client("cloudwatch")
        logs  = session.client("logs")

        # Fetch CloudTrail SSH/SSM events once for the whole region
        print(f"    {region}: fetching CloudTrail SSH/SSM events …")
        ct_access = get_cloudtrail_last_access(region)

        # ── PASS 1: collect raw instance data + build VPC/ENI maps ────────────
        raw_instances: list[dict] = []   # raw boto3 instance dicts
        all_vpc_ids:   set[str]  = set()
        all_eni_ids:   set[str]  = set()
        # sg_ids per instance collected here
        inst_sg_ids:  dict[str, list[str]] = {}   # iid → [sg_id, …]
        inst_eni_ids: dict[str, list[str]] = {}   # iid → [eni_id, …]

        paginator = ec2.get_paginator("describe_instances")
        pages = paginator.paginate(
            Filters=[{"Name": "instance-state-name", "Values": ["running", "stopped"]}]
        )
        for page in pages:
            for reservation in page.get("Reservations", []):
                for inst in reservation.get("Instances", []):
                    iid = inst["InstanceId"]
                    raw_instances.append(inst)
                    vpc_id = inst.get("VpcId", "")
                    if vpc_id:
                        all_vpc_ids.add(vpc_id)
                    sg_ids = [sg["GroupId"] for sg in inst.get("SecurityGroups", [])]
                    inst_sg_ids[iid] = sg_ids
                    eni_ids = [ni["NetworkInterfaceId"]
                               for ni in inst.get("NetworkInterfaces", [])]
                    inst_eni_ids[iid] = eni_ids
                    all_eni_ids.update(eni_ids)

        if not raw_instances:
            return result

        # ── VPC Flow Log groups (one lookup per region) ───────────────────────
        vpc_to_log_group = get_vpc_flow_log_groups(ec2, all_vpc_ids)
        flow_logs_available = bool(vpc_to_log_group)

        # ── Bulk Flow Log query per log group ─────────────────────────────────
        # Build: log_group → set of ENI IDs that belong to VPCs using that group
        lg_to_enis: dict[str, set[str]] = {}
        for inst in raw_instances:
            vpc_id = inst.get("VpcId", "")
            lg = vpc_to_log_group.get(vpc_id)
            if lg:
                for eni in inst_eni_ids.get(inst["InstanceId"], []):
                    lg_to_enis.setdefault(lg, set()).add(eni)

        # eni_id → formatted port string  (e.g. "22(1.2k), 443(890)")
        eni_port_map: dict[str, str] = {}
        if flow_logs_available:
            print(f"    {region}: querying VPC Flow Logs for port data …")
            for lg, eni_set in lg_to_enis.items():
                partial = query_flow_log_ports_bulk(logs, lg, eni_set)
                eni_port_map.update(partial)

        # ── PASS 2: build enriched instance records ───────────────────────────
        for inst in raw_instances:
            iid   = inst["InstanceId"]
            itype = inst.get("InstanceType", "unknown")
            state = inst["State"]["Name"]
            tags  = inst.get("Tags", [])
            name  = get_tag(tags, "Name", iid)
            owner = get_owner(tags)
            launch_time = inst.get("LaunchTime")
            platform = inst.get("Platform", "linux")
            vpc_id = inst.get("VpcId", "")

            # ── Security Group ports (always available) ───────────────────────
            sg_ports = get_sg_ports(ec2, inst_sg_ids.get(iid, []))

            # ── Flow Log ports (best-effort) ──────────────────────────────────
            enis = inst_eni_ids.get(iid, [])
            fl_parts: list[str] = []
            for eni in enis:
                if eni in eni_port_map:
                    fl_parts.append(eni_port_map[eni])
            fl_ports = ", ".join(fl_parts) if fl_parts else None
            # Combined port column
            if fl_ports:
                ports_str = f"FL: {fl_ports}"
            elif flow_logs_available:
                ports_str = f"SG: {sg_ports}  (FL: no traffic)"
            else:
                ports_str = f"SG: {sg_ports}"

            # ── CloudWatch metrics ────────────────────────────────────────────
            cpu_avg = None
            net_in_avg = None
            net_out_avg = None
            last_cw_activity = None
            last_network_active_day = None
            network_peak_packets = 0
            cloudtrail_last_access = ct_access.get(iid)

            if state == "running":
                cpu_avg                             = get_metric_avg(cw, iid, "CPUUtilization")
                net_in_avg                          = get_metric_avg(cw, iid, "NetworkIn")
                net_out_avg                         = get_metric_avg(cw, iid, "NetworkOut")
                last_cw_activity                    = get_last_datapoint_time(cw, iid)
                last_network_active_day, network_peak_packets = get_last_significant_network_day(cw, iid)

            # ── Best estimate of last human access ────────────────────────────
            last_human_access: datetime | None = None
            last_human_access_source: str = "none"
            access_details: str = ""
            candidates = []
            if cloudtrail_last_access:
                candidates.append((cloudtrail_last_access["time"], "CloudTrail SSH/SSM"))
            if last_network_active_day:
                candidates.append((last_network_active_day, "NetworkPacketsIn spike"))
            if candidates:
                last_human_access, last_human_access_source = max(candidates, key=lambda x: x[0])

            if cloudtrail_last_access:
                ct_ev = cloudtrail_last_access
                access_details = f"{ct_ev['event']} by {ct_ev['user']} from {ct_ev['source_ip']}"
            elif last_network_active_day:
                access_details = f"NetworkPacketsIn: {network_peak_packets:,} pkts on that day"

            # ── Idle / stopped age ────────────────────────────────────────────
            days_since_access: int | None = None
            if last_human_access:
                days_since_access = (datetime.now(timezone.utc) - last_human_access).days
            elif state == "running" and launch_time:
                days_since_access = (datetime.now(timezone.utc) - launch_time).days

            days_idle = None
            if state == "running":
                if last_cw_activity is None:
                    days_idle = (datetime.now(timezone.utc) - launch_time).days if launch_time else 0
                else:
                    days_idle = (datetime.now(timezone.utc) - last_cw_activity).days

            days_stopped = None
            if state == "stopped" and launch_time:
                days_stopped = (datetime.now(timezone.utc) - launch_time).days

            result["instances"].append({
                "id":           iid,
                "name":         name,
                "type":         itype,
                "state":        state,
                "launch_time":  launch_time,
                "platform":     platform,
                "owner":        owner,
                "tags":         tags,
                "cpu_avg_pct":  cpu_avg,
                "net_in_avg_bytes":  net_in_avg,
                "net_out_avg_bytes": net_out_avg,
                "last_cw_activity":  last_cw_activity,
                "last_network_active_day":   last_network_active_day,
                "network_peak_packets":      network_peak_packets,
                "cloudtrail_last_access":    cloudtrail_last_access,
                "last_human_access":         last_human_access,
                "last_human_access_source":  last_human_access_source,
                "access_details":            access_details,
                "days_since_access": days_since_access,
                "days_idle":    days_idle,
                "days_stopped": days_stopped,
                "sg_ports":     sg_ports,
                "fl_ports":     fl_ports,
                "ports_str":    ports_str,
            })

    except Exception as ex:
        result["error"] = str(ex)

    return result


# ── Categorisation ─────────────────────────────────────────────────────────────

def categorise(inst: dict) -> list[str]:
    """Return a list of concern labels for an instance."""
    labels = []
    state = inst["state"]
    cpu   = inst["cpu_avg_pct"]
    net_in  = inst["net_in_avg_bytes"] or 0
    net_out = inst["net_out_avg_bytes"] or 0
    days_idle = inst["days_idle"]
    days_stopped = inst["days_stopped"]

    if state == "running":
        if days_idle is not None and days_idle >= IDLE_DAYS_THRESHOLD:
            source = "no CW data" if inst["last_cw_activity"] is None else "last CW activity"
            labels.append(f"IDLE {days_idle}d ({source})")
        elif cpu is not None:
            if cpu < IDLE_CPU_THRESHOLD and (net_in + net_out) < IDLE_NETWORK_THRESHOLD * 86400:
                labels.append(f"IDLE  cpu={cpu:.1f}%")
            elif cpu < OVERALLOC_CPU_THRESHOLD:
                labels.append(f"OVER-ALLOCATED  cpu={cpu:.1f}%")

    elif state == "stopped":
        if days_stopped and days_stopped > 30:
            labels.append(f"STOPPED >{days_stopped}d")
        else:
            labels.append(f"STOPPED {days_stopped or '?'}d")

    return labels or ["OK"]


# ── Pretty printer ─────────────────────────────────────────────────────────────

def print_report(all_results: list[dict]):
    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    print("=" * 90)
    print(f"  EC2 AUDIT REPORT  –  {now_str}")
    print(f"  Lookback: {LOOKBACK_DAYS} days  |  Idle CPU threshold: {IDLE_CPU_THRESHOLD}%  "
          f"|  Over-alloc CPU threshold: {OVERALLOC_CPU_THRESHOLD}%")
    print(f"  Network 'real traffic' threshold: >{BACKGROUND_PACKETS_PER_DAY:,} packets/day  "
          f"|  CloudTrail lookback: {CLOUDTRAIL_LOOKBACK_DAYS} days")
    print("=" * 90)

    concern_buckets = {
        "IDLE": [],
        "OVER-ALLOCATED": [],
        "STOPPED": [],
        "OK": [],
    }

    total_instances = 0
    regions_with_errors = []

    for region_result in all_results:
        region = region_result["region"]
        if region_result["error"]:
            regions_with_errors.append(f"{region}: {region_result['error']}")
            continue
        for inst in region_result["instances"]:
            total_instances += 1
            labels = categorise(inst)
            primary = labels[0].split()[0]
            bucket = primary if primary in concern_buckets else "OK"
            concern_buckets[bucket].append((region, inst, labels))

    # ── Idle instances ──────────────────────────────────────────────────────────
    idle = concern_buckets["IDLE"]
    print(f"\n{'─'*90}")
    print(f"  🔴  IDLE / UNUSED INSTANCES  ({len(idle)} found)")
    print(f"{'─'*90}")
    if idle:
        for region, inst, labels in sorted(idle, key=lambda x: x[1].get("days_since_access") or 0, reverse=True):
            _print_instance_row(region, inst, labels)
    else:
        print("  None found.")

    # ── Over-allocated instances ────────────────────────────────────────────────
    overalloc = concern_buckets["OVER-ALLOCATED"]
    print(f"\n{'─'*90}")
    print(f"  🟡  OVER-ALLOCATED INSTANCES  ({len(overalloc)} found)")
    print(f"{'─'*90}")
    if overalloc:
        for region, inst, labels in sorted(overalloc, key=lambda x: x[1].get("cpu_avg_pct") or 0):
            _print_instance_row(region, inst, labels)
    else:
        print("  None found.")

    # ── Long-stopped instances ──────────────────────────────────────────────────
    stopped = concern_buckets["STOPPED"]
    long_stopped = [(r, i, l) for r, i, l in stopped if (i.get("days_stopped") or 0) > 30]
    print(f"\n{'─'*90}")
    print(f"  🟠  LONG-STOPPED INSTANCES >30d  ({len(long_stopped)} found)")
    print(f"{'─'*90}")
    if long_stopped:
        for region, inst, labels in sorted(long_stopped, key=lambda x: x[1].get("days_stopped") or 0, reverse=True):
            _print_instance_row(region, inst, labels)
    else:
        print("  None found.")

    # ── Summary ─────────────────────────────────────────────────────────────────
    print(f"\n{'─'*90}")
    print(f"  SUMMARY")
    print(f"{'─'*90}")
    print(f"  Total instances scanned : {total_instances}")
    print(f"  🔴 Idle / unused        : {len(idle)}")
    print(f"  🟡 Over-allocated       : {len(overalloc)}")
    print(f"  🟠 Long-stopped >30d    : {len(long_stopped)}")
    print(f"  ✅ Healthy (OK)         : {len(concern_buckets['OK'])}")

    if regions_with_errors:
        print(f"\n  ⚠️  Regions with errors (likely not enabled / no access):")
        for e in regions_with_errors[:10]:
            print(f"     • {e}")
        if len(regions_with_errors) > 10:
            print(f"     … and {len(regions_with_errors)-10} more")

    # ── Action items ─────────────────────────────────────────────────────────────
    action_targets = idle + overalloc + long_stopped
    if action_targets:
        print(f"\n{'─'*120}")
        print(f"  ACTION ITEMS  –  Instances to reach out about")
        print(f"{'─'*120}")
        print(f"  {'Instance ID':<22} {'Name':<22} {'Region':<14} {'When':<12} "
              f"{'Ports (SG=exposed, FL=actual)':<36} {'Access Detail':<38} Owner / Issue")
        print(f"  {'-'*21} {'-'*21} {'-'*13} {'-'*11} {'-'*35} {'-'*37} {'-'*30}")
        for region, inst, labels in action_targets:
            lha = inst.get("last_human_access")
            dsa = inst.get("days_since_access")
            det = inst.get("access_details", "") or "—"
            ports_str = inst.get("ports_str", "—")
            if lha:
                when_str = f"{dsa}d ago"
            elif dsa is not None:
                when_str = f">={dsa}d"
                det = "no access evidence"
            else:
                when_str = "unknown"
                det = "—"
            owner_issue = f"{inst['owner'][:18]}  {', '.join(labels)}"
            print(f"  {inst['id']:<22} {inst['name'][:21]:<22} {region:<14} "
                  f"{when_str:<12} {ports_str[:35]:<36} {det[:37]:<38} {owner_issue}")

    print(f"\n{'='*90}\n")


def _access_summary(inst: dict) -> str:
    """Short string describing last known human access (for action-items table)."""
    lha  = inst.get("last_human_access")
    dsa  = inst.get("days_since_access")
    det  = inst.get("access_details", "")
    if lha and det:
        return f"{dsa}d ago – {det}"
    if lha:
        return f"{dsa}d ago"
    if dsa is not None:
        return f">= {dsa}d (no access evidence)"
    return "unknown"


def _print_instance_row(region: str, inst: dict, labels: list[str]):
    cpu_str = f"{inst['cpu_avg_pct']:.1f}%" if inst["cpu_avg_pct"] is not None else "N/A"
    net_str = "N/A"
    if inst["net_in_avg_bytes"] is not None:
        total_kb = (inst["net_in_avg_bytes"] + (inst["net_out_avg_bytes"] or 0)) / 1024
        net_str = f"{total_kb:.0f} KB/day"
    last_act = format_age(inst["last_cw_activity"]) if inst["last_cw_activity"] else "no CW data"
    launch   = format_age(inst["launch_time"])

    # Human access lines
    lha  = inst.get("last_human_access")
    dsa  = inst.get("days_since_access")
    det  = inst.get("access_details", "")
    src  = inst.get("last_human_access_source", "none")

    if lha:
        human_when = f"{format_age(lha)}  [{src}]"
        human_what = det or "—"
    elif dsa is not None:
        human_when = f">= {dsa}d ago  (no SSH/SSM/network spike found)"
        human_what = "—"
    else:
        human_when = "unknown"
        human_what = "—"

    ports_str = inst.get("ports_str", "—")

    print(f"\n  {inst['id']}  [{inst['state'].upper()}]  {inst['type']}  |  {region}")
    print(f"    Name          : {inst['name']}")
    print(f"    Owner         : {inst['owner']}")
    print(f"    Launched      : {launch}")
    print(f"    CW activity   : {last_act}  |  CPU avg: {cpu_str}  |  Network avg: {net_str}")
    print(f"    Last access   : {human_when}")
    print(f"    Access detail : {human_what}")
    print(f"    Ports         : {ports_str}")
    print(f"    ⚠️   {', '.join(labels)}")


# ── Excel Export ────────────────────────────────────────────────────────────────

# Row fill colours (ARGB hex)
_FILL_IDLE     = PatternFill("solid", fgColor="FFCCCC")   # light red
_FILL_OVERALLOC= PatternFill("solid", fgColor="FFFFCC")   # light yellow
_FILL_STOPPED  = PatternFill("solid", fgColor="FFE5CC")   # light orange
_FILL_OK       = PatternFill("solid", fgColor="CCFFCC")   # light green
_FILL_HEADER   = PatternFill("solid", fgColor="1F4E79")   # dark navy

_FONT_HEADER   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_FONT_BODY     = Font(name="Calibri", size=10)
_FONT_BOLD     = Font(name="Calibri", bold=True, size=10)
_ALIGN_CENTER  = Alignment(horizontal="center", vertical="top", wrap_text=False)
_ALIGN_LEFT    = Alignment(horizontal="left",   vertical="top", wrap_text=True)

_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# ── column definition for the main sheet ────────────────────────────────────
_COLUMNS = [
    # (header text,           key/lambda,                           width)
    ("Instance ID",     lambda r, i: i["id"],                          18),
    ("Name",            lambda r, i: i["name"],                        24),
    ("Region",          lambda r, i: r,                                15),
    ("State",           lambda r, i: i["state"].upper(),               10),
    ("Type",            lambda r, i: i["type"],                        14),
    ("Platform",        lambda r, i: i.get("platform") or "linux",     10),
    ("Owner",           lambda r, i: i["owner"],                       22),
    ("Launched",        lambda r, i: (
        i["launch_time"].strftime("%Y-%m-%d") if i.get("launch_time") else "N/A"
    ),                                                                  12),
    ("Age (days)",      lambda r, i: (
        (datetime.now(timezone.utc) - i["launch_time"]).days
        if i.get("launch_time") else ""
    ),                                                                  10),
    ("Category",        lambda r, i: ", ".join(categorise(i)),         28),
    ("CPU Avg %",       lambda r, i: (
        round(i["cpu_avg_pct"], 1) if i["cpu_avg_pct"] is not None else ""
    ),                                                                  10),
    ("Net In KB/day",   lambda r, i: (
        round(i["net_in_avg_bytes"] / 1024, 1)
        if i.get("net_in_avg_bytes") is not None else ""
    ),                                                                  13),
    ("Net Out KB/day",  lambda r, i: (
        round(i["net_out_avg_bytes"] / 1024, 1)
        if i.get("net_out_avg_bytes") is not None else ""
    ),                                                                  13),
    ("Days Idle/Stopped", lambda r, i: (
        i.get("days_idle") if i["state"] == "running"
        else i.get("days_stopped") or ""
    ),                                                                  16),
    ("Last Human Access", lambda r, i: (
        i["last_human_access"].strftime("%Y-%m-%d")
        if i.get("last_human_access") else ""
    ),                                                                  16),
    ("Days Since Access", lambda r, i: i.get("days_since_access") or "",  16),
    ("Access Source",   lambda r, i: i.get("last_human_access_source") or "",  22),
    ("Access Detail",   lambda r, i: i.get("access_details") or "",    38),
    ("Ports (SG/exposed)", lambda r, i: i.get("sg_ports") or "",       26),
    ("Ports (FL/actual)",  lambda r, i: i.get("fl_ports") or "",       26),
]


def _row_fill(labels: list[str]) -> PatternFill:
    primary = labels[0].split()[0]
    if primary == "IDLE":         return _FILL_IDLE
    if primary == "OVER-ALLOCATED": return _FILL_OVERALLOC
    if primary == "STOPPED":      return _FILL_STOPPED
    return _FILL_OK


def _write_sheet(ws, rows: list[tuple], title: str):
    """Write header + data rows to worksheet `ws`.

    rows = list of (region_str, inst_dict)
    """
    ws.title = title

    # ── Header ───────────────────────────────────────────────────────────────
    for col_idx, (hdr, _, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font      = _FONT_HEADER
        cell.fill      = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border    = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions   # set after data is written (see below)

    # ── Data rows ────────────────────────────────────────────────────────────
    for row_idx, (region, inst) in enumerate(rows, start=2):
        labels = categorise(inst)
        fill   = _row_fill(labels)

        for col_idx, (hdr, getter, _) in enumerate(_COLUMNS, start=1):
            try:
                value = getter(region, inst)
            except Exception:
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = fill
            cell.font      = _FONT_BODY
            cell.border    = _BORDER
            cell.alignment = _ALIGN_LEFT

        # row height – comfortable for wrapped text
        ws.row_dimensions[row_idx].height = 18

    # Update auto-filter to cover all data
    if ws.max_row >= 2:
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(_COLUMNS))}{ws.max_row}"
        )


def _write_summary_sheet(ws, all_results: list[dict]):
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16

    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    rows_meta = [
        ("EC2 Audit Report", ""),
        ("Generated",        generated),
        ("Lookback (days)",  LOOKBACK_DAYS),
        ("Idle CPU threshold",f"< {IDLE_CPU_THRESHOLD}%"),
        ("Over-alloc CPU threshold", f"< {OVERALLOC_CPU_THRESHOLD}%"),
        ("Background packet threshold", f"> {BACKGROUND_PACKETS_PER_DAY:,} pkt/day"),
        ("CloudTrail lookback", f"{CLOUDTRAIL_LOOKBACK_DAYS} days"),
        ("", ""),
        ("Category", "Count"),
    ]

    # Tally categories
    counts = {"IDLE": 0, "OVER-ALLOCATED": 0, "STOPPED": 0, "OK": 0}
    total  = 0
    regions_seen = set()
    for rr in all_results:
        if rr.get("error"):
            continue
        regions_seen.add(rr["region"])
        for inst in rr["instances"]:
            total += 1
            labels  = categorise(inst)
            primary = labels[0].split()[0]
            bucket  = primary if primary in counts else "OK"
            counts[bucket] += 1

    long_stopped = sum(
        1 for rr in all_results if not rr.get("error")
        for inst in rr["instances"]
        if inst["state"] == "stopped" and (inst.get("days_stopped") or 0) > 30
    )

    for label, val in rows_meta:
        r = ws.max_row + 1 if ws.max_row else 1
        ws.append([label, val])

    ws.append(["🔴 Idle / unused",     counts["IDLE"]])
    ws.append(["🟡 Over-allocated",    counts["OVER-ALLOCATED"]])
    ws.append(["🟠 Long-stopped >30d", long_stopped])
    ws.append(["✅ Healthy (OK)",       counts["OK"]])
    ws.append(["Total instances",       total])
    ws.append(["Regions scanned",       len(regions_seen)])

    # Style first column bold, header row navy
    for row in ws.iter_rows():
        for cell in row:
            cell.font = _FONT_BODY
            cell.alignment = Alignment(horizontal="left", vertical="top")
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    # Category header row
    for cell in ws["9:9"]:
        cell.font = _FONT_BOLD
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER


# ── AWS Services Sheet ──────────────────────────────────────────────────────────

_SVC_FILLS: dict[str, PatternFill] = {
    "RDS":          PatternFill("solid", fgColor="D6E4F7"),   # soft blue
    "Lambda":       PatternFill("solid", fgColor="FFF2CC"),   # soft yellow
    "S3":           PatternFill("solid", fgColor="E2F0D9"),   # soft green
    "ELB/APP":      PatternFill("solid", fgColor="FCE4D6"),   # soft orange
    "ELB/NET":      PatternFill("solid", fgColor="FCE4D6"),
    "ELB/GAT":      PatternFill("solid", fgColor="FCE4D6"),
    "ElastiCache":  PatternFill("solid", fgColor="E2CFEF"),   # soft purple
    "DynamoDB":     PatternFill("solid", fgColor="FDEBD0"),   # soft peach
    "SQS":          PatternFill("solid", fgColor="D5F5E3"),   # soft mint
    "ECS Cluster":  PatternFill("solid", fgColor="EAF2FB"),   # lighter blue
    "ECS Service":  PatternFill("solid", fgColor="D6EAF8"),
    "CloudFront":   PatternFill("solid", fgColor="F9EBEA"),   # soft pink
}
_SVC_FILL_DEFAULT = PatternFill("solid", fgColor="F2F2F2")

_SVC_COLUMNS = [
    ("Service",              "service",  15),
    ("Name / ID",            "name",     36),
    ("Region",               "region",   15),
    ("Type / Config",        "stype",    30),
    ("State / Status",       "state",    14),
    ("Metric 1",             "m1l",      24),
    (f"Value ({LOOKBACK_DAYS}d)", "m1v", 16),
    ("Metric 2",             "m2l",      24),
    (f"Value ({LOOKBACK_DAYS}d) ", "m2v", 16),
    ("Notes",                "notes",    46),
    ("Created / Modified",   "created",  18),
]


def _write_aws_services_sheet(ws, service_rows: list[dict]):
    """Write the AWS Native Services sheet."""
    ws.title = "AWS Services"

    for ci, (hdr, _, w) in enumerate(_SVC_COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font      = _FONT_HEADER
        cell.fill      = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border    = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"

    sorted_rows = sorted(
        service_rows,
        key=lambda r: (r.get("service", ""), r.get("region", ""), r.get("name", "").lower()),
    )

    for ri, row in enumerate(sorted_rows, start=2):
        svc_key = row.get("service", "")
        fill    = _SVC_FILLS.get(svc_key, _SVC_FILL_DEFAULT)
        for ci, (_, key, _) in enumerate(_SVC_COLUMNS, start=1):
            val = row.get(key, "")
            if isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d")
            elif isinstance(val, str) and len(val) > 32_700:
                val = val[:32_697] + "…"
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = fill
            cell.font      = _FONT_BODY
            cell.border    = _BORDER
            cell.alignment = _ALIGN_LEFT
        ws.row_dimensions[ri].height = 16

    if ws.max_row >= 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_SVC_COLUMNS))}{ws.max_row}"


# ── Subscribed Services Sheet ──────────────────────────────────────────────────

_SUB_FILLS: dict[str, PatternFill] = {
    # Reserved Instances
    "EC2 Reserved RI":         PatternFill("solid", fgColor="D6E4F7"),  # blue
    "RDS Reserved RI":         PatternFill("solid", fgColor="BDD7EE"),  # darker blue
    "ElastiCache Reserved RI": PatternFill("solid", fgColor="9DC3E6"),  # even darker blue
    # Savings Plans
    "Savings Plan":            PatternFill("solid", fgColor="E2F0D9"),  # green
    # Security - enabled
    "GuardDuty":               PatternFill("solid", fgColor="D5F5E3"),  # mint
    "Security Hub":            PatternFill("solid", fgColor="D5F5E3"),
    "Inspector v2":            PatternFill("solid", fgColor="D5F5E3"),
    "Macie":                   PatternFill("solid", fgColor="D5F5E3"),
    "AWS Config":              PatternFill("solid", fgColor="D5F5E3"),
    # Other global
    "Shield Advanced":         PatternFill("solid", fgColor="F2DCDB"),  # soft red
    "Support Plan":            PatternFill("solid", fgColor="FFF2CC"),  # yellow
    "Marketplace":             PatternFill("solid", fgColor="FCE4D6"),  # peach
}
_SUB_FILL_NOT_ENABLED = PatternFill("solid", fgColor="F2F2F2")   # grey  = not enabled
_SUB_FILL_DEFAULT     = PatternFill("solid", fgColor="FFFFFF")

_SUB_COLUMNS = [
    ("Service",              "service",  22),
    ("Name / ID",            "name",     36),
    ("Region",               "region",   15),
    ("Type / Config",        "stype",    30),
    ("State / Status",       "state",    16),
    ("Detail 1",             "m1l",      26),
    ("Value",                "m1v",      22),
    ("Detail 2",             "m2l",      22),
    ("Value ",               "m2v",      16),
    ("Notes",                "notes",    44),
    ("Created / Start",      "created",  18),
]

# States that indicate "not enabled / not subscribed"
_NOT_ENABLED_STATES = {
    "NOT ENABLED", "NOT SUBSCRIBED", "NOT RECORDING", "DISABLED",
    "NO SPEND", "disabled",
}


def _write_subscribed_services_sheet(ws, sub_rows: list[dict]):
    """Write the Subscribed / Activated Services sheet."""
    ws.title = "Subscribed Services"

    for ci, (hdr, _, w) in enumerate(_SUB_COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font      = _FONT_HEADER
        cell.fill      = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border    = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"

    # Sort: service type → region → name
    sorted_rows = sorted(
        sub_rows,
        key=lambda r: (r.get("service",""), r.get("region",""), r.get("name","").lower()),
    )

    for ri, row in enumerate(sorted_rows, start=2):
        state   = str(row.get("state",""))
        svc_key = row.get("service","")
        if state.upper() in _NOT_ENABLED_STATES:
            fill = _SUB_FILL_NOT_ENABLED
        else:
            fill = _SUB_FILLS.get(svc_key, _SUB_FILL_DEFAULT)

        for ci, (_, key, _) in enumerate(_SUB_COLUMNS, start=1):
            val = row.get(key, "")
            if isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d")
            elif isinstance(val, str) and len(val) > 32_700:
                val = val[:32_697] + "…"
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = fill
            cell.font      = _FONT_BODY
            cell.border    = _BORDER
            cell.alignment = _ALIGN_LEFT
        ws.row_dimensions[ri].height = 16

    if ws.max_row >= 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_SUB_COLUMNS))}{ws.max_row}"


# ── Marketplace Sheet ──────────────────────────────────────────────────────────

_FILL_MKT_LICENSE_AVAIL  = PatternFill("solid", fgColor="E2F0D9")   # green
_FILL_MKT_LICENSE_OTHER  = PatternFill("solid", fgColor="F2F2F2")   # grey
_FILL_MKT_SPEND_DATA     = PatternFill("solid", fgColor="D6E4F7")   # blue
_FILL_MKT_SPEND_ZERO     = PatternFill("solid", fgColor="FFFFFF")   # white
_FILL_MKT_SUBHDR         = PatternFill("solid", fgColor="2E75B6")   # mid-blue
_FONT_MKT_SUBHDR         = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_FONT_MKT_MONEY          = Font(name="Calibri", size=10, color="1F4E79")


def _write_marketplace_sheet(ws, marketplace_data: dict):
    """
    Write the Marketplace sheet with two sections:
      A. Licensed Products (License Manager)
      B. Monthly Spend History (Cost Explorer – last 14 months)
    """
    ws.title = "Marketplace"

    licenses     = marketplace_data.get("licenses",     [])
    spend_rows   = marketplace_data.get("spend_rows",   [])
    spend_months = marketplace_data.get("spend_months", [])

    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 42
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 22

    row = 1

    # ── Title ─────────────────────────────────────────────────────────────────
    title_cell = ws.cell(row=row, column=1, value="AWS Marketplace & License Manager")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    ws.cell(row=row, column=2, value=f"Generated: {generated}").font = _FONT_BODY
    row += 2

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION A – Licensed Products
    # ══════════════════════════════════════════════════════════════════════════
    sub_hdr_a = ws.cell(row=row, column=1,
                        value=f"SECTION A — Licensed Products ({len(licenses)} licenses)")
    sub_hdr_a.font = _FONT_MKT_SUBHDR
    sub_hdr_a.fill = _FILL_MKT_SUBHDR
    for col in range(1, 9):
        ws.cell(row=row, column=col).fill = _FILL_MKT_SUBHDR
    row += 1

    # Column headers
    lic_headers = [
        ("Product Name",   42),
        ("Status",         14),
        ("Issuer",         22),
        ("Product SKU",    22),
        ("Valid From",     14),
        ("Valid To",       14),
        ("Entitlements",   52),
        ("License ARN",    46),
    ]
    lic_keys = ["product","status","issuer","sku","valid_from","valid_to","entitlements","license_arn"]
    for ci, (hdr, w) in enumerate(lic_headers, start=1):
        cell = ws.cell(row=row, column=ci, value=hdr)
        cell.font      = _FONT_HEADER
        cell.fill      = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border    = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{row + 1}"
    row += 1

    if licenses:
        for lic in sorted(licenses, key=lambda l: l.get("product","").lower()):
            status = lic.get("status","")
            fill = _FILL_MKT_LICENSE_AVAIL if status == "AVAILABLE" else _FILL_MKT_LICENSE_OTHER
            for ci, key in enumerate(lic_keys, start=1):
                val = lic.get(key, "")
                cell = ws.cell(row=row, column=ci, value=val)
                cell.fill      = fill
                cell.font      = _FONT_BODY
                cell.border    = _BORDER
                cell.alignment = _ALIGN_LEFT
            ws.row_dimensions[row].height = 16
            row += 1
    else:
        ws.cell(row=row, column=1, value="No licenses found in License Manager.").font = _FONT_BODY
        row += 1

    row += 2  # blank spacer

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION B – Monthly Spend
    # ══════════════════════════════════════════════════════════════════════════
    sub_hdr_b = ws.cell(row=row, column=1,
                        value="SECTION B — Marketplace Monthly Spend (Cost Explorer)")
    sub_hdr_b.font = _FONT_MKT_SUBHDR
    sub_hdr_b.fill = _FILL_MKT_SUBHDR
    n_spend_cols = 3 + len(spend_months) + 1   # RecordType + Operation + months + Total
    for col in range(1, n_spend_cols + 1):
        ws.cell(row=row, column=col).fill = _FILL_MKT_SUBHDR
    row += 1

    if not spend_months:
        ws.cell(row=row, column=1,
                value="No Marketplace charges found in Cost Explorer history.").font = _FONT_BODY
    else:
        # Header row: Record Type | Operation | <month>... | Total
        spend_hdr_cols = ["Record Type", "Operation / Product"] + spend_months + ["Total (USD)"]
        spend_col_widths = [26, 40] + [12] * len(spend_months) + [14]
        for ci, (hdr, w) in enumerate(zip(spend_hdr_cols, spend_col_widths), start=1):
            cell = ws.cell(row=row, column=ci, value=hdr)
            cell.font      = _FONT_HEADER
            cell.fill      = _FILL_HEADER
            cell.alignment = _ALIGN_CENTER
            cell.border    = _BORDER
            ws.column_dimensions[get_column_letter(ci)].width = w
        row += 1

        has_any_spend = any(r["total"] > 0 for r in spend_rows)

        if has_any_spend:
            for sr in spend_rows:
                total = sr["total"]
                fill  = _FILL_MKT_SPEND_DATA if total > 0 else _FILL_MKT_SPEND_ZERO
                ws.cell(row=row, column=1, value=sr["record_type"]).fill = fill
                ws.cell(row=row, column=2, value=sr["operation"]).fill   = fill
                for mi, month in enumerate(spend_months, start=3):
                    amt = sr["monthly"].get(month, 0.0)
                    cell = ws.cell(row=row, column=mi,
                                   value=round(amt, 2) if amt > 0 else "")
                    cell.fill = fill
                    cell.font = _FONT_MKT_MONEY if amt > 0 else _FONT_BODY
                    cell.border = _BORDER
                    cell.alignment = _ALIGN_LEFT
                total_cell = ws.cell(row=row, column=3 + len(spend_months),
                                     value=round(total, 2) if total > 0 else "")
                total_cell.fill = fill
                total_cell.font = _FONT_MKT_MONEY if total > 0 else _FONT_BODY
                total_cell.border = _BORDER
                # Apply fill/font/border to first two cols
                for ci in range(1, 3):
                    ws.cell(row=row, column=ci).font   = _FONT_BODY
                    ws.cell(row=row, column=ci).border = _BORDER
                    ws.cell(row=row, column=ci).alignment = _ALIGN_LEFT
                ws.row_dimensions[row].height = 16
                row += 1
        else:
            ws.cell(row=row, column=1,
                    value="$0.00 — No Marketplace charges in any recorded month.").font = _FONT_BODY
            row += 1

        # Grand total row
        row += 1
        grand = sum(r["total"] for r in spend_rows)
        ws.cell(row=row, column=2, value="Grand Total").font = _FONT_BOLD
        gt_cell = ws.cell(row=row, column=3 + len(spend_months), value=round(grand, 2))
        gt_cell.font = Font(name="Calibri", bold=True, size=10, color="1F4E79")
        gt_cell.fill = PatternFill("solid", fgColor="BDD7EE")
        gt_cell.border = _BORDER


def write_excel_report(all_results: list[dict], service_rows: list[dict],
                       sub_rows: list[dict], marketplace_data: dict, path: str):
    """
    Write a colour-coded Excel workbook to `path` with seven sheets:
      1. Summary            – counts & configuration
      2. All Instances      – every EC2 found
      3. Action Items       – IDLE, OVER-ALLOCATED, STOPPED >30d
      4. Legend             – colour key & column descriptions
      5. AWS Services       – usage stats for all other AWS native services
      6. Subscribed Services– RIs, Savings Plans, security services, Marketplace summary
      7. Marketplace        – License Manager licenses + CE monthly spend history
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    _write_summary_sheet(wb.active, all_results)

    # ── Build flat row lists ──────────────────────────────────────────────────
    all_rows:    list[tuple] = []
    action_rows: list[tuple] = []

    for rr in all_results:
        if rr.get("error"):
            continue
        for inst in rr["instances"]:
            all_rows.append((rr["region"], inst))
            labels  = categorise(inst)
            primary = labels[0].split()[0]
            is_long_stopped = (
                primary == "STOPPED"
                and (inst.get("days_stopped") or 0) > 30
            )
            if primary in ("IDLE", "OVER-ALLOCATED") or is_long_stopped:
                action_rows.append((rr["region"], inst))

    # Sort action items: IDLE first (by days_since_access desc), then
    # OVER-ALLOCATED (by cpu asc), then STOPPED (by days_stopped desc)
    def _sort_key(t):
        _, inst = t
        labels  = categorise(inst)
        primary = labels[0].split()[0]
        order   = {"IDLE": 0, "OVER-ALLOCATED": 1, "STOPPED": 2}.get(primary, 3)
        if primary == "IDLE":
            return (order, -(inst.get("days_since_access") or 0))
        if primary == "OVER-ALLOCATED":
            return (order, inst.get("cpu_avg_pct") or 0)
        return (order, -(inst.get("days_stopped") or 0))

    action_rows.sort(key=_sort_key)
    all_rows.sort(key=lambda t: (t[0], t[1]["name"]))  # sort by region then name

    # ── Sheet 2: All Instances ────────────────────────────────────────────────
    ws_all = wb.create_sheet()
    _write_sheet(ws_all, all_rows, "All Instances")

    # ── Sheet 3: Action Items ─────────────────────────────────────────────────
    ws_action = wb.create_sheet()
    _write_sheet(ws_action, action_rows, "Action Items")

    # ── Legend sheet ─────────────────────────────────────────────────────────
    ws_leg = wb.create_sheet("Legend")
    ws_leg.column_dimensions["A"].width = 22
    ws_leg.column_dimensions["B"].width = 50
    legend_rows = [
        ("Colour",         "Meaning"),
        ("🔴 Light red",   "IDLE – running but no meaningful activity"),
        ("🟡 Light yellow","OVER-ALLOCATED – running but CPU consistently < 20%"),
        ("🟠 Light orange","STOPPED > 30 days – consider terminating"),
        ("🟢 Light green", "OK – healthy / recently active"),
        ("",               ""),
        ("Column",         "Description"),
        ("Days Idle/Stopped", "For running: days since last CloudWatch data point. "
                              "For stopped: days since launch time."),
        ("Last Human Access", "Most recent date a real human accessed the instance "
                              "(CloudTrail SSH/SSM OR NetworkPacketsIn spike > "
                              f"{BACKGROUND_PACKETS_PER_DAY:,} pkts/day)."),
        ("Days Since Access", "Days since last human access (or instance age if "
                              "no access signal found)."),
        ("Ports (SG/exposed)","Inbound ports open from the public internet "
                              "(0.0.0.0/0 or ::/0 sources in Security Groups)."),
        ("Ports (FL/actual)", "Destination ports that received ACCEPT traffic "
                              "according to VPC Flow Logs (blank if Flow Logs not enabled)."),
    ]
    for i, (a, b) in enumerate(legend_rows, start=1):
        ws_leg.cell(row=i, column=1, value=a).font = _FONT_BOLD if i in (1, 7) else _FONT_BODY
        ws_leg.cell(row=i, column=2, value=b).font = _FONT_BODY
        ws_leg.cell(row=i, column=2).alignment = Alignment(horizontal="left", wrap_text=True)
        ws_leg.row_dimensions[i].height = 30

    # ── Sheet 5: AWS Native Services ─────────────────────────────────────────
    ws_svc = wb.create_sheet()
    _write_aws_services_sheet(ws_svc, service_rows)

    # ── Sheet 6: Subscribed Services ──────────────────────────────────────────
    ws_sub = wb.create_sheet()
    _write_subscribed_services_sheet(ws_sub, sub_rows)

    # ── Sheet 7: Marketplace ───────────────────────────────────────────────────
    ws_mkt = wb.create_sheet()
    _write_marketplace_sheet(ws_mkt, marketplace_data)

    wb.save(path)
    print(f"  Excel report saved to : {path}")


# ── AWS Native Services Collection ────────────────────────────────────────────

def _cw_stat(cw_client, namespace: str, metric: str, dims: list,
             stat: str = "Average", days: int = LOOKBACK_DAYS) -> float | None:
    """Return daily-granularity Sum or Average of a CW metric over `days` days."""
    end   = datetime.now(timezone.utc)
    start = end - timedelta(days=days)
    try:
        resp = cw_client.get_metric_statistics(
            Namespace=namespace, MetricName=metric, Dimensions=dims,
            StartTime=start, EndTime=end, Period=86400, Statistics=[stat],
        )
        dps = resp.get("Datapoints", [])
        if not dps:
            return None
        return (sum(d[stat] for d in dps)
                if stat == "Sum"
                else sum(d[stat] for d in dps) / len(dps))
    except Exception:
        return None


def _svc_rec(service: str, name: str, region: str, stype: str, state: str,
             m1l: str, m1v, m2l: str, m2v, notes: str = "", created=None) -> dict:
    return dict(service=service, name=name, region=region, stype=stype,
                state=state, m1l=m1l, m1v=m1v, m2l=m2l, m2v=m2v,
                notes=notes, created=created)


def _collect_rds(session, region: str, cw) -> list[dict]:
    rows = []
    try:
        rds = session.client("rds")
        for page in rds.get_paginator("describe_db_instances").paginate():
            for db in page["DBInstances"]:
                iid  = db["DBInstanceIdentifier"]
                dims = [{"Name": "DBInstanceIdentifier", "Value": iid}]
                cpu  = _cw_stat(cw, "AWS/RDS", "CPUUtilization",      dims)
                conn = _cw_stat(cw, "AWS/RDS", "DatabaseConnections", dims)
                rows.append(_svc_rec(
                    "RDS", iid, region,
                    f"{db.get('DBInstanceClass','')} / {db.get('Engine','')} {db.get('EngineVersion','')}",
                    db.get("DBInstanceStatus", ""),
                    "CPU Avg %",     round(cpu,  1) if cpu  is not None else "",
                    "DB Conns Avg",  round(conn, 1) if conn is not None else "",
                    f"Storage: {db.get('AllocatedStorage','')} GB  MultiAZ: {db.get('MultiAZ','')}",
                    db.get("InstanceCreateTime"),
                ))
    except Exception:
        pass
    return rows


def _collect_lambda(session, region: str, cw) -> list[dict]:
    rows = []
    try:
        lam = session.client("lambda")
        for page in lam.get_paginator("list_functions").paginate():
            for fn in page["Functions"]:
                name = fn["FunctionName"]
                dims = [{"Name": "FunctionName", "Value": name}]
                inv  = _cw_stat(cw, "AWS/Lambda", "Invocations", dims, stat="Sum")
                err  = _cw_stat(cw, "AWS/Lambda", "Errors",      dims, stat="Sum")
                dur  = _cw_stat(cw, "AWS/Lambda", "Duration",    dims, stat="Average")
                rows.append(_svc_rec(
                    "Lambda", name, region,
                    f"{fn.get('Runtime','')}  {fn.get('MemorySize','')} MB",
                    fn.get("State", "Active"),
                    f"Invocations ({LOOKBACK_DAYS}d)", int(inv) if inv is not None else 0,
                    "Errors",  int(err) if err is not None else 0,
                    f"Avg duration: {round(dur, 0)} ms" if dur else "",
                    fn.get("LastModified"),
                ))
    except Exception:
        pass
    return rows


def _collect_elb(session, region: str, cw) -> list[dict]:
    rows = []
    try:
        elb = session.client("elbv2")
        for page in elb.get_paginator("describe_load_balancers").paginate():
            for lb in page["LoadBalancers"]:
                name    = lb["LoadBalancerName"]
                arn     = lb["LoadBalancerArn"]
                lb_type = lb.get("Type", "application")
                # CW dimension = everything after the first segment of the ARN path suffix
                arn_suffix = arn.split(":")[-1]           # e.g. "loadbalancer/app/my-lb/abc123"
                lb_dim = "/".join(arn_suffix.split("/")[1:])  # "app/my-lb/abc123"
                ns   = ("AWS/ApplicationELB" if lb_type == "application"
                        else "AWS/NetworkELB")
                dims = [{"Name": "LoadBalancer", "Value": lb_dim}]
                reqs = _cw_stat(cw, ns, "RequestCount",          dims, stat="Sum")
                conn = _cw_stat(cw, ns, "ActiveConnectionCount", dims, stat="Average")
                rows.append(_svc_rec(
                    f"ELB/{lb_type[:3].upper()}", name, region,
                    lb.get("Scheme", ""),
                    lb.get("State", {}).get("Code", ""),
                    f"Requests ({LOOKBACK_DAYS}d)", int(reqs) if reqs is not None else "",
                    "Active Conns Avg", round(conn, 0) if conn is not None else "",
                    f"DNS: {lb.get('DNSName','')[:60]}",
                    lb.get("CreatedTime"),
                ))
    except Exception:
        pass
    return rows


def _collect_elasticache(session, region: str, cw) -> list[dict]:
    rows = []
    try:
        ec_cli = session.client("elasticache")
        for page in ec_cli.get_paginator("describe_cache_clusters").paginate(ShowCacheNodeInfo=True):
            for cl in page["CacheClusters"]:
                cid  = cl["CacheClusterId"]
                dims = [{"Name": "CacheClusterId", "Value": cid}]
                cpu  = _cw_stat(cw, "AWS/ElastiCache", "CPUUtilization",  dims)
                conn = _cw_stat(cw, "AWS/ElastiCache", "CurrConnections", dims)
                rows.append(_svc_rec(
                    "ElastiCache", cid, region,
                    f"{cl.get('CacheNodeType','')} / {cl.get('Engine','')} {cl.get('EngineVersion','')}",
                    cl.get("CacheClusterStatus", ""),
                    "CPU Avg %",  round(cpu,  1) if cpu  is not None else "",
                    "Conns Avg",  round(conn, 0) if conn is not None else "",
                    f"Nodes: {cl.get('NumCacheNodes','')}",
                    cl.get("CacheClusterCreateTime"),
                ))
    except Exception:
        pass
    return rows


def _collect_dynamodb(session, region: str, cw) -> list[dict]:
    rows = []
    try:
        ddb = session.client("dynamodb")
        for page in ddb.get_paginator("list_tables").paginate():
            for tname in page["TableNames"]:
                try:
                    desc = ddb.describe_table(TableName=tname)["Table"]
                except Exception:
                    desc = {}
                dims = [{"Name": "TableName", "Value": tname}]
                rcu  = _cw_stat(cw, "AWS/DynamoDB", "ConsumedReadCapacityUnits",  dims, stat="Sum")
                wcu  = _cw_stat(cw, "AWS/DynamoDB", "ConsumedWriteCapacityUnits", dims, stat="Sum")
                items   = desc.get("ItemCount", "")
                size_mb = (round(desc.get("TableSizeBytes", 0) / 1024**2, 1)
                           if desc.get("TableSizeBytes") else "")
                rows.append(_svc_rec(
                    "DynamoDB", tname, region,
                    desc.get("BillingModeSummary", {}).get("BillingMode", "PROVISIONED"),
                    desc.get("TableStatus", ""),
                    f"Read CU ({LOOKBACK_DAYS}d)",  int(rcu) if rcu is not None else 0,
                    f"Write CU ({LOOKBACK_DAYS}d)", int(wcu) if wcu is not None else 0,
                    f"Items: {items}  Size: {size_mb} MB",
                    desc.get("CreationDateTime"),
                ))
    except Exception:
        pass
    return rows


def _collect_sqs(session, region: str, cw) -> list[dict]:
    rows = []
    try:
        sqs = session.client("sqs")
        next_token = None
        while True:
            kwargs: dict = {}
            if next_token:
                kwargs["NextToken"] = next_token
            resp = sqs.list_queues(**kwargs)
            for url in resp.get("QueueUrls", []):
                qname = url.split("/")[-1]
                try:
                    attrs = sqs.get_queue_attributes(
                        QueueUrl=url,
                        AttributeNames=["ApproximateNumberOfMessages", "CreatedTimestamp"],
                    )["Attributes"]
                except Exception:
                    attrs = {}
                dims  = [{"Name": "QueueName", "Value": qname}]
                sent  = _cw_stat(cw, "AWS/SQS", "NumberOfMessagesSent", dims, stat="Sum")
                depth = int(attrs.get("ApproximateNumberOfMessages", 0) or 0)
                ts    = attrs.get("CreatedTimestamp")
                created_dt = (datetime.fromtimestamp(int(ts), tz=timezone.utc) if ts else None)
                rows.append(_svc_rec(
                    "SQS", qname, region, "Queue", "active",
                    f"Msgs Sent ({LOOKBACK_DAYS}d)", int(sent) if sent is not None else 0,
                    "Queue Depth", depth,
                    "", created_dt,
                ))
            next_token = resp.get("NextToken")
            if not next_token:
                break
    except Exception:
        pass
    return rows


def _collect_ecs(session, region: str, cw) -> list[dict]:
    rows = []
    try:
        ecs = session.client("ecs")
        cluster_arns: list[str] = []
        for page in ecs.get_paginator("list_clusters").paginate():
            cluster_arns.extend(page["clusterArns"])
        if not cluster_arns:
            return rows
        clusters = ecs.describe_clusters(clusters=cluster_arns[:100])["clusters"]
        for cl in clusters:
            cl_name  = cl["clusterName"]
            svc_arns: list[str] = []
            for page in ecs.get_paginator("list_services").paginate(cluster=cl_name):
                svc_arns.extend(page["serviceArns"])
            if not svc_arns:
                rows.append(_svc_rec(
                    "ECS Cluster", cl_name, region, "", cl.get("status", ""),
                    "Running Tasks",   cl.get("runningTasksCount", ""),
                    "Active Services", cl.get("activeServicesCount", ""),
                    "",
                ))
                continue
            all_svcs: list[dict] = []
            for i in range(0, min(len(svc_arns), 100), 10):
                try:
                    all_svcs.extend(ecs.describe_services(
                        cluster=cl_name, services=svc_arns[i:i+10]
                    ).get("services", []))
                except Exception:
                    pass
            for sd in all_svcs:
                sname = sd["serviceName"]
                dims  = [{"Name": "ClusterName", "Value": cl_name},
                         {"Name": "ServiceName",  "Value": sname}]
                cpu = _cw_stat(cw, "AWS/ECS", "CPUUtilization",    dims)
                mem = _cw_stat(cw, "AWS/ECS", "MemoryUtilization", dims)
                rows.append(_svc_rec(
                    "ECS Service", sname, region,
                    f"Cluster: {cl_name}  {sd.get('launchType','')}",
                    sd.get("status", ""),
                    "CPU Avg %", round(cpu, 1) if cpu is not None else "",
                    "Mem Avg %", round(mem, 1) if mem is not None else "",
                    f"Running: {sd.get('runningCount','')}  Desired: {sd.get('desiredCount','')}",
                    sd.get("createdAt"),
                ))
    except Exception:
        pass
    return rows


def _collect_s3_global() -> list[dict]:
    """Collect S3 bucket list + CloudWatch storage metrics (published daily to us-east-1)."""
    rows = []
    try:
        s3  = boto3.client("s3",          region_name="us-east-1")
        cw  = boto3.client("cloudwatch",  region_name="us-east-1")
        buckets = s3.list_buckets().get("Buckets", [])
        for bkt in buckets:
            bname    = bkt["Name"]
            dims_std = [{"Name": "BucketName",  "Value": bname},
                        {"Name": "StorageType", "Value": "StandardStorage"}]
            dims_all = [{"Name": "BucketName",  "Value": bname},
                        {"Name": "StorageType", "Value": "AllStorageTypes"}]
            size_b  = _cw_stat(cw, "AWS/S3", "BucketSizeBytes", dims_std, stat="Average", days=3)
            obj_cnt = _cw_stat(cw, "AWS/S3", "NumberOfObjects",  dims_all, stat="Average", days=3)
            size_gb = round(size_b / 1024**3, 3) if size_b else ""
            try:
                loc     = s3.get_bucket_location(Bucket=bname)
                bregion = loc.get("LocationConstraint") or "us-east-1"
            except Exception:
                bregion = "unknown"
            rows.append(_svc_rec(
                "S3", bname, bregion, "Bucket", "active",
                "Size (GB)",  size_gb,
                "Objects",    int(obj_cnt) if obj_cnt else "",
                "", bkt.get("CreationDate"),
            ))
    except Exception:
        pass
    return rows


def _collect_cloudfront_global() -> list[dict]:
    """Collect CloudFront distribution stats (global service; CW metrics in us-east-1)."""
    rows = []
    try:
        cf = boto3.client("cloudfront", region_name="us-east-1")
        cw = boto3.client("cloudwatch",  region_name="us-east-1")
        for page in cf.get_paginator("list_distributions").paginate():
            for dist in page.get("DistributionList", {}).get("Items", []):
                did     = dist["Id"]
                domain  = dist.get("DomainName", "")
                origins = ", ".join(
                    o.get("DomainName", "")
                    for o in dist.get("Origins", {}).get("Items", [])
                )
                dims = [{"Name": "DistributionId", "Value": did},
                        {"Name": "Region",          "Value": "Global"}]
                reqs = _cw_stat(cw, "AWS/CloudFront", "Requests",      dims, stat="Sum")
                errs = _cw_stat(cw, "AWS/CloudFront", "5xxErrorRate",  dims, stat="Average")
                rows.append(_svc_rec(
                    "CloudFront", did, "global",
                    dist.get("Comment") or domain,
                    dist.get("Status", ""),
                    f"Requests ({LOOKBACK_DAYS}d)",  int(reqs) if reqs is not None else "",
                    "5xx Error Rate %",  round(errs, 2) if errs is not None else "",
                    f"Domain: {domain}  Origins: {origins[:60]}",
                    None,
                ))
    except Exception:
        pass
    return rows


def collect_all_services(regions: list[str]) -> list[dict]:
    """
    Scan all regions for AWS native services (RDS, Lambda, ELB, ECS, ElastiCache,
    DynamoDB, SQS) plus global services (S3, CloudFront).
    Returns a flat list of _svc_rec dicts ready for the Excel sheet.
    """
    print("\nScanning AWS native services across all regions …")
    all_rows: list[dict] = []

    def _scan_region(region: str) -> list[dict]:
        rrows: list[dict] = []
        try:
            session = boto3.session.Session(region_name=region)
            cw = session.client("cloudwatch")
            rrows.extend(_collect_rds(session, region, cw))
            rrows.extend(_collect_lambda(session, region, cw))
            rrows.extend(_collect_elb(session, region, cw))
            rrows.extend(_collect_elasticache(session, region, cw))
            rrows.extend(_collect_dynamodb(session, region, cw))
            rrows.extend(_collect_sqs(session, region, cw))
            rrows.extend(_collect_ecs(session, region, cw))
        except Exception:
            pass
        return rrows

    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = {pool.submit(_scan_region, r): r for r in regions}
        done = 0
        for future in as_completed(futures):
            done += 1
            r     = futures[future]
            rrows = future.result()
            all_rows.extend(rrows)
            status = f"{len(rrows)} resource(s)" if rrows else "—"
            print(f"  [{done:>2}/{len(regions)}] {r:<22} {status}")

    print("  Scanning global services (S3, CloudFront) …")
    all_rows.extend(_collect_s3_global())
    all_rows.extend(_collect_cloudfront_global())
    print(f"  Total AWS service resources found: {len(all_rows)}\n")
    return all_rows


# ── Subscribed / Activated Services Collection ────────────────────────────────

def _collect_reserved_instances(region: str) -> list[dict]:
    """Active Reserved Instances for EC2, RDS, ElastiCache, Redshift."""
    rows = []

    # EC2
    try:
        ec2 = boto3.client("ec2", region_name=region)
        resp = ec2.describe_reserved_instances(
            Filters=[{"Name": "state", "Values": ["active"]}]
        )
        for ri in resp.get("ReservedInstances", []):
            end_dt = ri.get("End")
            days_left = (end_dt - datetime.now(timezone.utc)).days if end_dt else None
            rows.append(_svc_rec(
                "EC2 Reserved RI", ri["ReservedInstancesId"], region,
                f"{ri.get('InstanceType','')} × {ri.get('InstanceCount',1)}",
                ri.get("State", ""),
                "Offering",      f"{ri.get('OfferingType','')} / {ri.get('OfferingClass','')}",
                "Days Remaining", days_left if days_left is not None else "",
                f"Platform: {ri.get('ProductDescription','')}  Tenancy: {ri.get('InstanceTenancy','')}",
                ri.get("Start"),
            ))
    except Exception:
        pass

    # RDS
    try:
        rds = boto3.client("rds", region_name=region)
        for page in rds.get_paginator("describe_reserved_db_instances").paginate():
            for ri in page["ReservedDBInstances"]:
                if ri.get("State") != "active":
                    continue
                start = ri.get("StartTime")
                dur   = ri.get("Duration", 0)
                end_dt = (start + timedelta(seconds=dur)) if start else None
                days_left = (end_dt - datetime.now(timezone.utc)).days if end_dt else None
                rows.append(_svc_rec(
                    "RDS Reserved RI", ri["ReservedDBInstanceId"], region,
                    f"{ri.get('DBInstanceClass','')} × {ri.get('DBInstanceCount',1)}",
                    ri.get("State", ""),
                    "Offering",       ri.get("OfferingType",""),
                    "Days Remaining", days_left if days_left is not None else "",
                    f"Engine: {ri.get('ProductDescription','')}  MultiAZ: {ri.get('MultiAZ','')}",
                    start,
                ))
    except Exception:
        pass

    # ElastiCache
    try:
        ec_cli = boto3.client("elasticache", region_name=region)
        for page in ec_cli.get_paginator("describe_reserved_cache_nodes").paginate():
            for ri in page["ReservedCacheNodes"]:
                if ri.get("State") != "active":
                    continue
                start = ri.get("StartTime")
                dur   = ri.get("Duration", 0)
                end_dt = (start + timedelta(seconds=dur)) if start else None
                days_left = (end_dt - datetime.now(timezone.utc)).days if end_dt else None
                rows.append(_svc_rec(
                    "ElastiCache Reserved RI", ri["ReservedCacheNodeId"], region,
                    f"{ri.get('CacheNodeType','')} × {ri.get('CacheNodeCount',1)}",
                    ri.get("State", ""),
                    "Offering",       ri.get("OfferingType",""),
                    "Days Remaining", days_left if days_left is not None else "",
                    f"Engine: {ri.get('ProductDescription','')}",
                    start,
                ))
    except Exception:
        pass

    return rows


def _collect_security_services(region: str, account_id: str) -> list[dict]:
    """
    Check which security/compliance services are enabled in a region.
    Adds one row per service showing ENABLED / NOT ENABLED.
    """
    rows = []

    # ── GuardDuty ──────────────────────────────────────────────────────────────
    try:
        gd  = boto3.client("guardduty", region_name=region)
        det_ids = gd.list_detectors().get("DetectorIds", [])
        if det_ids:
            det = gd.get_detector(DetectorId=det_ids[0])
            rows.append(_svc_rec(
                "GuardDuty", det_ids[0], region,
                det.get("FindingPublishingFrequency",""),
                det.get("Status","ENABLED"),
                "Status",           det.get("Status","ENABLED"),
                "Finding Frequency", det.get("FindingPublishingFrequency",""),
                "",
                det.get("CreatedAt"),
            ))
        else:
            rows.append(_svc_rec("GuardDuty", "—", region, "", "NOT ENABLED",
                                  "Status", "NOT ENABLED", "", "", ""))
    except Exception:
        pass

    # ── Security Hub ───────────────────────────────────────────────────────────
    try:
        sh  = boto3.client("securityhub", region_name=region)
        hub = sh.describe_hub()
        rows.append(_svc_rec(
            "Security Hub",
            (hub.get("HubArn","").split("/")[-1] or "hub"),
            region, "",
            "ENABLED",
            "Status",              "ENABLED",
            "Auto-Enable Controls", str(hub.get("AutoEnableControls","")),
            "",
            hub.get("SubscribedAt"),
        ))
    except Exception as ex:
        if any(k in str(ex) for k in ("InvalidAccessException", "not subscribed", "not enabled")):
            rows.append(_svc_rec("Security Hub", "—", region, "", "NOT ENABLED",
                                  "Status", "NOT ENABLED", "", "", ""))

    # ── Amazon Inspector v2 ────────────────────────────────────────────────────
    try:
        insp = boto3.client("inspector2", region_name=region)
        resp = insp.batch_get_account_status(accountIds=[account_id])
        for acct in resp.get("accounts", []):
            state = acct.get("state", {}).get("status", "DISABLED")
            ec2_s = acct.get("resourceState", {}).get("ec2",  {}).get("status","")
            ecr_s = acct.get("resourceState", {}).get("ecr",  {}).get("status","")
            rows.append(_svc_rec(
                "Inspector v2", acct.get("accountId",""), region, "", state,
                "Status",     state,
                "EC2 / ECR",  f"{ec2_s} / {ecr_s}",
                "",
            ))
    except Exception as ex:
        if any(k in str(ex) for k in ("ValidationException","AccessDeniedException")):
            pass   # service not available in region; skip silently
        else:
            rows.append(_svc_rec("Inspector v2", "—", region, "", "NOT ENABLED",
                                  "Status", "NOT ENABLED", "", "", ""))

    # ── Amazon Macie ───────────────────────────────────────────────────────────
    try:
        macie = boto3.client("macie2", region_name=region)
        sess  = macie.get_macie_session()
        rows.append(_svc_rec(
            "Macie", "", region, "", sess.get("status",""),
            "Status",             sess.get("status",""),
            "Finding Frequency",  sess.get("findingPublishingFrequency",""),
            "",
            sess.get("createdAt"),
        ))
    except Exception as ex:
        if any(k in str(ex) for k in ("MacieClientError","not enabled","AccessDeniedException")):
            rows.append(_svc_rec("Macie", "—", region, "", "NOT ENABLED",
                                  "Status", "NOT ENABLED", "", "", ""))

    # ── AWS Config ─────────────────────────────────────────────────────────────
    try:
        cfg = boto3.client("config", region_name=region)
        statuses = cfg.describe_configuration_recorder_status().get("ConfigurationRecordersStatus", [])
        if statuses:
            rec = statuses[0]
            rows.append(_svc_rec(
                "AWS Config", rec.get("name","default"), region, "",
                "RECORDING" if rec.get("recording") else "NOT RECORDING",
                "Recording",    str(rec.get("recording",False)),
                "Last Status",  rec.get("lastStatus",""),
                f"Last error: {rec.get('lastErrorCode','none')}",
            ))
        else:
            rows.append(_svc_rec("AWS Config", "—", region, "", "NOT ENABLED",
                                  "Status", "NOT ENABLED", "", "", ""))
    except Exception:
        pass

    return rows


def _collect_savings_plans() -> list[dict]:
    """Active Savings Plans (global)."""
    rows = []
    try:
        sp = boto3.client("savingsplans", region_name="us-east-1")
        next_token = None
        while True:
            kwargs: dict = {"states": ["active"]}
            if next_token:
                kwargs["nextToken"] = next_token
            resp = sp.describe_savings_plans(**kwargs)
            for plan in resp.get("savingsPlans", []):
                end_raw = plan.get("end")
                if isinstance(end_raw, str):
                    end_dt: datetime | None = datetime.fromisoformat(
                        end_raw.replace("Z", "+00:00")
                    )
                elif isinstance(end_raw, datetime):
                    end_dt = end_raw
                else:
                    end_dt = None
                days_left = (end_dt - datetime.now(timezone.utc)).days if end_dt else None
                term_yr = int(plan.get("termDurationInSeconds", 0)) // 86400 // 365
                rows.append(_svc_rec(
                    "Savings Plan", plan.get("savingsPlanId",""), "global",
                    f"{plan.get('savingsPlanType','')} / {plan.get('paymentOption','')} / {term_yr}yr",
                    plan.get("state",""),
                    "Commitment $/hr",  plan.get("commitment",""),
                    "Days Remaining",   days_left if days_left is not None else "",
                    f"Region: {plan.get('region','any')}  Family: {plan.get('ec2InstanceFamily','any')}",
                    plan.get("start"),
                ))
            next_token = resp.get("nextToken")
            if not next_token:
                break
    except Exception:
        pass
    return rows


def _collect_shield() -> list[dict]:
    """AWS Shield Advanced subscription status (global)."""
    rows = []
    try:
        shield = boto3.client("shield", region_name="us-east-1")
        sub = shield.describe_subscription()["Subscription"]
        end_dt = sub.get("EndTime")
        days_left = (end_dt - datetime.now(timezone.utc)).days if end_dt else None
        rows.append(_svc_rec(
            "Shield Advanced", "subscription", "global",
            f"AutoRenew: {sub.get('AutoRenew','')}",
            "ENABLED",
            "Proactive Engagement", sub.get("ProactiveEngagementStatus",""),
            "Days Remaining",       days_left if days_left is not None else "",
            f"Protected resources limits: {len(sub.get('Limits',[]))} defined",
            sub.get("StartTime"),
        ))
    except Exception as ex:
        if any(k in str(ex) for k in ("ResourceNotFoundException","SubscriptionNotFoundException")):
            rows.append(_svc_rec(
                "Shield Advanced", "—", "global", "Standard (free)", "NOT SUBSCRIBED",
                "Status", "Standard only (free tier)", "", "", "",
            ))
    return rows


def _collect_support_plan() -> list[dict]:
    """Detect the active AWS Support plan tier."""
    rows = []
    try:
        support = boto3.client("support", region_name="us-east-1")
        sevs  = support.describe_severity_levels(language="en")["severityLevels"]
        codes = {s["code"] for s in sevs}
        if "critical" in codes:
            tier = "Enterprise"
        elif "urgent" in codes:
            tier = "Business"
        elif "high" in codes:
            tier = "Developer"
        else:
            tier = "Basic"
        rows.append(_svc_rec(
            "Support Plan", "account", "global", tier, "ACTIVE",
            "Tier",              tier,
            "Severity Levels",   ", ".join(sorted(codes)),
            "",
        ))
    except Exception as ex:
        if "SubscriptionRequiredException" in str(ex):
            rows.append(_svc_rec(
                "Support Plan", "account", "global", "Basic", "ACTIVE",
                "Tier", "Basic (free)", "Note", "No premium Support API access", "",
            ))
    return rows


def _collect_marketplace_subscriptions() -> list[dict]:
    """
    Lightweight summary row for the Subscribed Services sheet.
    Full detail lives in the dedicated Marketplace sheet (collect_marketplace_data).
    """
    rows = []
    try:
        ce    = boto3.client("ce", region_name="us-east-1")
        end   = datetime.now(timezone.utc).date()
        start = end - timedelta(days=30)
        total_resp = ce.get_cost_and_usage(
            TimePeriod={"Start": str(start), "End": str(end)},
            Granularity="MONTHLY",
            Filter={"Dimensions": {"Key": "SERVICE", "Values": ["AWS Marketplace"]}},
            Metrics=["UnblendedCost"],
        )
        total_cost = sum(
            float(p.get("Total", {}).get("UnblendedCost", {}).get("Amount", 0))
            for p in total_resp.get("ResultsByTime", [])
        )
        label = "NO SPEND" if total_cost <= 0 else "ACTIVE"
        rows.append(_svc_rec(
            "Marketplace", "summary", "global",
            "See 'Marketplace' sheet for full detail",
            label,
            "Cost (30d USD)", f"${total_cost:.2f}",
            "Licenses", "(see Marketplace sheet)",
            "",
        ))
    except Exception:
        pass
    return rows


# ── Marketplace Sheet Data Collection ─────────────────────────────────────────

def collect_marketplace_data() -> dict:
    """
    Collect full Marketplace data for its own Excel sheet:
      - License Manager received licenses (free + Bedrock + BYOL + paid)
      - Cost Explorer monthly spend (last 14 months) by product/operation
    Returns dict with keys: 'licenses', 'spend_rows', 'spend_months'
    """
    print("  Collecting Marketplace / License Manager data …")
    result: dict = {"licenses": [], "spend_rows": [], "spend_months": []}

    # ── License Manager ───────────────────────────────────────────────────────
    try:
        lm   = boto3.client("license-manager", region_name="us-east-1")
        resp = lm.list_received_licenses(MaxResults=100)
        lics = resp.get("Licenses", [])
        while resp.get("NextToken"):
            resp = lm.list_received_licenses(MaxResults=100, NextToken=resp["NextToken"])
            lics.extend(resp.get("Licenses", []))
        for lic in lics:
            validity = lic.get("Validity", {})
            ent_str  = ", ".join(
                f"{e.get('Name','')}="
                f"{e.get('MaxCount', e.get('Value', '∞'))}"
                for e in lic.get("Entitlements", [])[:6]
            )
            result["licenses"].append({
                "product":     lic.get("ProductName", ""),
                "status":      lic.get("Status", ""),
                "issuer":      lic.get("Issuer", {}).get("Name", ""),
                "sku":         lic.get("ProductSKU", ""),
                "license_arn": lic.get("LicenseArn", ""),
                "valid_from":  validity.get("Begin", ""),
                "valid_to":    validity.get("End", ""),
                "entitlements": ent_str,
            })
    except Exception as ex:
        print(f"    License Manager: {ex}")

    # ── Cost Explorer – monthly spend by product/operation (last 14 months) ──
    try:
        ce    = boto3.client("ce", region_name="us-east-1")
        today = datetime.now(timezone.utc).date()
        # Start on first of month 13 months ago (CE requires first-of-month for >14m)
        start_month = (today.replace(day=1) - timedelta(days=395)).replace(day=1)

        resp = ce.get_cost_and_usage(
            TimePeriod={"Start": str(start_month), "End": str(today)},
            Granularity="MONTHLY",
            Filter={"Dimensions": {"Key": "SERVICE", "Values": ["AWS Marketplace"]}},
            GroupBy=[
                {"Type": "DIMENSION", "Key": "RECORD_TYPE"},
                {"Type": "DIMENSION", "Key": "OPERATION"},
            ],
            Metrics=["UnblendedCost"],
        )

        from collections import defaultdict
        months_set: set[str]  = set()
        spend_map:  dict      = defaultdict(lambda: defaultdict(float))

        for period in resp.get("ResultsByTime", []):
            month = period["TimePeriod"]["Start"][:7]
            months_set.add(month)
            for grp in period.get("Groups", []):
                keys = tuple(grp.get("Keys", ["?", "?"]))
                amt  = float(grp.get("Metrics", {}).get("UnblendedCost", {}).get("Amount", 0))
                spend_map[keys][month] += amt

        result["spend_months"] = sorted(months_set)
        for keys, monthly in spend_map.items():
            result["spend_rows"].append({
                "record_type": keys[0] if len(keys) > 0 else "",
                "operation":   keys[1] if len(keys) > 1 else "",
                "monthly":     dict(monthly),
                "total":       sum(monthly.values()),
            })
        result["spend_rows"].sort(key=lambda r: -r["total"])

    except Exception as ex:
        print(f"    Cost Explorer (Marketplace): {ex}")

    return result


def collect_subscribed_services(regions: list[str]) -> list[dict]:
    """
    Collect AWS subscription data across all regions:
      - EC2 / RDS / ElastiCache Reserved Instances
      - Security services per region (GuardDuty, Security Hub, Inspector v2, Macie, Config)
      - Global: Savings Plans, Shield Advanced, Support plan, Marketplace spend
    """
    print("\nScanning subscribed / activated services …")
    all_rows: list[dict] = []

    # Get account ID once for Inspector v2 calls
    try:
        account_id = boto3.client("sts").get_caller_identity()["Account"]
    except Exception:
        account_id = ""

    def _scan_region(region: str) -> list[dict]:
        rrows: list[dict] = []
        rrows.extend(_collect_reserved_instances(region))
        rrows.extend(_collect_security_services(region, account_id))
        return rrows

    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = {pool.submit(_scan_region, r): r for r in regions}
        done = 0
        for future in as_completed(futures):
            done += 1
            r     = futures[future]
            rrows = future.result()
            all_rows.extend(rrows)
            status = f"{len(rrows)} record(s)" if rrows else "—"
            print(f"  [{done:>2}/{len(regions)}] {r:<22} {status}")

    print("  Scanning global subscriptions …")
    all_rows.extend(_collect_savings_plans())
    all_rows.extend(_collect_shield())
    all_rows.extend(_collect_support_plan())
    all_rows.extend(_collect_marketplace_subscriptions())
    print(f"  Total subscription records: {len(all_rows)}\n")
    return all_rows


# ── Main ────────────────────────────────────────────────────────────────────────

def main():
    print("Discovering regions …")
    try:
        ec2_meta = boto3.client("ec2", region_name="us-east-1")
        regions_resp = ec2_meta.describe_regions(AllRegions=False)
        regions = [r["RegionName"] for r in regions_resp["Regions"]
                   if r["RegionName"] not in SKIP_REGIONS]
    except Exception as ex:
        print(f"Could not list regions: {ex}")
        regions = ["us-east-1", "us-west-2", "eu-west-1", "ap-southeast-1"]

    print(f"Scanning {len(regions)} regions in parallel (this may take ~2-3 min) …\n")

    all_results = []
    with ThreadPoolExecutor(max_workers=10) as pool:
        futures = {pool.submit(audit_region, r): r for r in regions}
        done = 0
        for future in as_completed(futures):
            done += 1
            r = futures[future]
            res = future.result()
            count = len(res["instances"])
            err   = res["error"]
            status = f"{count} instances" if not err else f"ERROR: {err[:60]}"
            print(f"  [{done:>2}/{len(regions)}] {r:<20} {status}")
            all_results.append(res)

    print()
    print_report(all_results)

    # ── Collect AWS native service stats ──────────────────────────────────────
    service_rows = collect_all_services(regions)

    # ── Collect subscribed / activated services ────────────────────────────────
    sub_rows = collect_subscribed_services(regions)

    # ── Collect Marketplace / License Manager data ─────────────────────────────
    marketplace_data = collect_marketplace_data()

    output_path = "/tmp/ec2_audit_results.json"
    with open(output_path, "w") as f:
        def serial(obj):
            if isinstance(obj, datetime):
                return obj.isoformat()
            raise TypeError(f"Not serializable: {type(obj)}")
        json.dump(all_results, f, indent=2, default=serial)
    print(f"  Raw JSON results saved to: {output_path}\n")

    excel_path = "/tmp/ec2_audit_results.xlsx"
    write_excel_report(all_results, service_rows, sub_rows, marketplace_data, excel_path)


if __name__ == "__main__":
    main()
