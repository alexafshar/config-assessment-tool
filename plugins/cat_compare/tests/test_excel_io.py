from openpyxl import Workbook
from openpyxl import load_workbook

from compare_tool.excel_io import (
    inspect_summary_formula_cache,
    summary_cache_is_available,
    summary_missing_cache_is_supported,
)
from compare_tool.summary import create_summary_workbooks, read_summary_values_df


def test_summary_cache_available_when_summary_has_values_only(tmp_path):
    workbook_path = tmp_path / "values_only.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    ws["A2"] = "Total"
    ws["B2"] = 42
    wb.save(workbook_path)

    info = inspect_summary_formula_cache(str(workbook_path))

    assert info["summary_exists"] is True
    assert info["formula_cells"] == 0
    assert info["missing_cached_formula_cells"] == 0
    assert summary_cache_is_available(str(workbook_path)) is True


def test_summary_cache_missing_when_formula_has_no_cached_value(tmp_path):
    workbook_path = tmp_path / "formula_without_cache.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    ws["A2"] = "Total"
    ws["B2"] = "=SUM(1,2)"
    wb.save(workbook_path)

    info = inspect_summary_formula_cache(str(workbook_path))

    assert info["summary_exists"] is True
    assert info["formula_cells"] == 1
    assert info["missing_cached_formula_cells"] == 1
    assert info["missing_cached_coordinates"] == ["B2"]
    assert summary_cache_is_available(str(workbook_path)) is False
    assert summary_missing_cache_is_supported(str(workbook_path)) is False


def test_summary_cache_unavailable_when_summary_sheet_is_missing(tmp_path):
    workbook_path = tmp_path / "missing_summary.xlsx"
    wb = Workbook()
    wb.active.title = "Analysis"
    wb.save(workbook_path)

    info = inspect_summary_formula_cache(str(workbook_path))

    assert info["summary_exists"] is False
    assert summary_cache_is_available(str(workbook_path)) is False


def test_supported_cat_summary_formulas_can_be_synthesized(tmp_path):
    previous_path = tmp_path / "previous.xlsx"
    current_path = tmp_path / "current.xlsx"
    previous_sum_path = tmp_path / "previous_sum.xlsx"
    current_sum_path = tmp_path / "current_sum.xlsx"

    for path in (previous_path, current_path):
        wb = Workbook()
        summary = wb.active
        summary.title = "Summary"
        summary.append([None, "bronze", "silver", "gold", "platinum"])
        summary.append([
            "# of Apps",
            '=COUNTIF(INDEX(Analysis!$1:$1048576,0,MATCH("OverallAssessment",Analysis!$1:$1,0)), B1)',
            '=COUNTIF(INDEX(Analysis!$1:$1048576,0,MATCH("OverallAssessment",Analysis!$1:$1,0)), C1)',
            '=COUNTIF(INDEX(Analysis!$1:$1048576,0,MATCH("OverallAssessment",Analysis!$1:$1,0)), D1)',
            '=COUNTIF(INDEX(Analysis!$1:$1048576,0,MATCH("OverallAssessment",Analysis!$1:$1,0)), E1)',
        ])
        summary.append([
            "% of Apps",
            "=ROUND(B2/(COUNTA(Analysis!$C:$C)-1)*100, 1)",
            "=ROUND(C2/(COUNTA(Analysis!$C:$C)-1)*100, 1)",
            "=ROUND(D2/(COUNTA(Analysis!$C:$C)-1)*100, 1)",
            "=ROUND(E2/(COUNTA(Analysis!$C:$C)-1)*100, 1)",
        ])

        analysis = wb.create_sheet("Analysis")
        analysis.append(["controller", "id", "name", "OverallAssessment"])
        analysis.append(["ctrl", 1, "App A", "bronze"])
        analysis.append(["ctrl", 2, "App B", "silver"])
        analysis.append(["ctrl", 3, "App C", "gold"])
        analysis.append(["ctrl", 4, "App D", "gold"])
        analysis.append(["ctrl", 5, "App E", "platinum"])
        wb.save(path)

    assert summary_missing_cache_is_supported(str(current_path)) is True

    current_summary_df = read_summary_values_df(str(current_path))
    assert current_summary_df.columns.tolist() == [
        "Unnamed: 0",
        "bronze",
        "silver",
        "gold",
        "platinum",
    ]
    assert current_summary_df.loc[0, "bronze"] == 1
    assert current_summary_df.loc[1, "gold"] == 40

    create_summary_workbooks(
        str(previous_path),
        str(current_path),
        str(previous_sum_path),
        str(current_sum_path),
    )

    wb = load_workbook(current_sum_path, data_only=True)
    try:
        ws = wb["Summary"]
        assert [ws["B2"].value, ws["C2"].value, ws["D2"].value, ws["E2"].value] == [1, 1, 2, 1]
        assert [ws["B3"].value, ws["C3"].value, ws["D3"].value, ws["E3"].value] == [20, 20, 40, 20]
    finally:
        wb.close()
