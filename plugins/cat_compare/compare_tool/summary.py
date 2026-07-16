"""
summary.py
----------
This module handles the generation and processing of summary data.

Purpose:
- Extracts and processes summary data from Excel files.
- Generates summary reports for comparisons.

Key Features:
- Reads the "Summary" sheet from Excel files.
- Processes metrics and generates summary data for use in comparisons.
"""

# compare_tool/summary.py

import logging
import re
from typing import Any
from copy import copy
from openpyxl.styles import PatternFill


import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from pathlib import Path

import pandas as pd

from .excel_io import get_key_column  # if used

logger = logging.getLogger(__name__)


def _build_summary_formula_context(wb_formula, wb_values):
    context = {
        "overall_counts": {},
        "analysis_col_c_count": 0,
    }

    if "Analysis" not in wb_values.sheetnames:
        return context

    ws = wb_values["Analysis"]
    headers = {
        str(cell.value or "").strip(): cell.column
        for cell in next(ws.iter_rows(min_row=1, max_row=1), [])
    }

    overall_col = headers.get("OverallAssessment")
    if overall_col:
        counts = {}
        for row in ws.iter_rows(min_row=2, min_col=overall_col, max_col=overall_col):
            value = row[0].value
            if value is None:
                continue
            key = str(value).strip().lower()
            if key:
                counts[key] = counts.get(key, 0) + 1
        context["overall_counts"] = counts

    col_c_values = [
        cell.value
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3)
        for cell in row
        if cell.value not in (None, "")
    ]
    context["analysis_col_c_count"] = len(col_c_values)
    return context


def _summary_formula_value(ws_output, ws_formula, cell, context):
    formula = str(cell.value or "")
    formula_upper = formula.upper()

    if "COUNTIF(" in formula_upper and "OVERALLASSESSMENT" in formula_upper:
        header = ws_formula.cell(row=1, column=cell.column).value
        return context["overall_counts"].get(str(header or "").strip().lower(), 0)

    if formula_upper.startswith("=ROUND(") and "COUNTA(ANALYSIS!" in formula_upper:
        match = re.search(r"=ROUND\(([A-Z]+)(\d+)/", formula_upper)
        if not match:
            return None
        count_cell = f"{match.group(1)}{match.group(2)}"
        count_value = ws_output[count_cell].value or 0
        app_count = context["analysis_col_c_count"]
        if not app_count:
            return 0
        return round((float(count_value) / app_count) * 100, 1)

    return None


def _summary_values_workbook(source_file_path):
    wb_formula = load_workbook(source_file_path, data_only=False)
    wb_values = load_workbook(source_file_path, data_only=True)
    try:
        if "Summary" not in wb_formula.sheetnames or "Summary" not in wb_values.sheetnames:
            logging.error("'Summary' sheet is missing in %s.", source_file_path)
            return None

        ws_formula = wb_formula["Summary"]
        ws_values = wb_values["Summary"]
        context = _build_summary_formula_context(wb_formula, wb_values)

        wb_output = openpyxl.Workbook()
        ws_output = wb_output.active
        ws_output.title = "Summary"

        for row in ws_formula.iter_rows():
            for cell in row:
                value = ws_values[cell.coordinate].value
                formula_text = cell.value
                is_formula = cell.data_type == "f" or (
                    isinstance(formula_text, str) and formula_text.startswith("=")
                )
                if is_formula and value is None:
                    value = _summary_formula_value(ws_output, ws_formula, cell, context)
                    if value is None:
                        logging.warning(
                            "Could not calculate Summary formula %s in %s: %s",
                            cell.coordinate,
                            source_file_path,
                            formula_text,
                        )
                ws_output.cell(row=cell.row, column=cell.column, value=value)

        return wb_output
    finally:
        wb_formula.close()
        wb_values.close()


def read_summary_values_df(source_file_path):
    wb_output = _summary_values_workbook(source_file_path)
    if wb_output is None:
        return pd.DataFrame()

    ws_output = wb_output["Summary"]
    rows = list(ws_output.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()

    headers = [
        f"Unnamed: {idx}" if value is None else value
        for idx, value in enumerate(rows[0])
    ]
    return pd.DataFrame(rows[1:], columns=headers)


def _copy_summary_values(source_file_path, output_path):
    wb_output = _summary_values_workbook(source_file_path)
    if wb_output is not None:
        wb_output.save(output_path)


# Function to create summary workbooks
def create_summary_workbooks(previous_file_path, current_file_path, previous_sum_path, current_sum_path):
    try:
        _copy_summary_values(previous_file_path, previous_sum_path)
        _copy_summary_values(current_file_path, current_sum_path)

    except Exception as e:
        logging.error(f"Error in create_summary_workbooks: {e}", exc_info=True)
        raise


# Function to compare 'Summary' sheet and save to a new workbook
def compare_files_summary(previous_sum_path, current_sum_path, comparison_sum_path):
    try:
        # Load the previous_sum and current_sum Excel files
        wb_previous = load_workbook(previous_sum_path, data_only=True)
        wb_current = load_workbook(current_sum_path, data_only=True)
        wb_output = openpyxl.Workbook()

        ws_previous = wb_previous['Summary']
        ws_current = wb_current['Summary']
        ws_output = wb_output.active
        ws_output.title = 'Summary'
        
        logging.debug(f"Processing sheet: 'Summary'")
        
        compare_summary(ws_previous, ws_current, ws_output)

        # Save the workbook after all modifications have been completed
        wb_output.save(comparison_sum_path)
        logging.debug(f"Summary comparison saved to: {comparison_sum_path}")

    except Exception as e:
        logging.error(f"Error in compare_files_summary: {e}", exc_info=True)
        raise


# Function to compare summaries of both sheets
def compare_summary(ws_previous, ws_current, ws_output):
    from openpyxl.styles import PatternFill

    # Define fill styles
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for row in ws_previous.iter_rows(min_row=1, min_col=1, max_col=ws_previous.max_column, max_row=ws_previous.max_row):
        for cell in row:
            prev_cell = ws_previous.cell(row=cell.row, column=cell.column)
            curr_cell = ws_current.cell(row=cell.row, column=cell.column)
            output_cell = ws_output.cell(row=cell.row, column=cell.column)

            prev_value = prev_cell.value
            curr_value = curr_cell.value

            if prev_value is None:
                prev_value = ''
            if curr_value is None:
                curr_value = ''

            logging.debug(f"Comparing cell ({cell.row},{cell.column}): Previous Value: {prev_value}, Current Value: {curr_value}")

            if prev_value != curr_value:
                if isinstance(prev_value, (int, float)) and isinstance(curr_value, (int, float)):
                    if curr_value > prev_value:
                        output_cell.fill = green_fill
                    else:
                        output_cell.fill = red_fill
                    output_cell.value = f"{prev_value} → {curr_value}"
                else:
                    output_cell.fill = red_fill
                    output_cell.value = f"{prev_value} → {curr_value}"
            else:
                output_cell.value = prev_value

            logging.debug(f"Cell ({cell.row},{cell.column}) updated to: {output_cell.value}")


def eval_formula(formula: str, context: dict[str, Any]) -> Any:
    """
    Your existing helper to evaluate formulas in the summary comparison, if any.
    """
    # --- BEGIN: copy body from old eval_formula -------------------
    # def eval_formula(...):
    #     ...
    # --- END -------------------------------------------------------


def copy_summary_to_result(comparison_sum_path: str, output_file_path: str) -> None:
    """
    Copy the 'Summary' sheet from comparison_sum_path into the result workbook
    at output_file_path.

    - If the result workbook doesn't exist yet, create a blank one first
    - Insert Summary as the FIRST sheet (index 0)
    - Copy values + cell styles
    - Apply Bronze/Silver/Gold/Platinum header colours
    """
    try:
        comp_path = Path(comparison_sum_path)
        out_path = Path(output_file_path)

        if not comp_path.exists():
            logging.warning(
                "Comparison summary workbook %s does not exist; nothing to copy.",
                comp_path,
            )
            return

        # Ensure the result workbook exists
        if not out_path.exists():
            logging.warning(
                "Result workbook %s not found; creating a new workbook before "
                "copying summary.",
                out_path,
            )
            out_path.parent.mkdir(parents=True, exist_ok=True)
            wb_new = Workbook()
            wb_new.save(out_path)

        # Load both workbooks
        wb_comparison_sum = load_workbook(comp_path)
        wb_output = load_workbook(out_path)

        if "Summary" not in wb_comparison_sum.sheetnames:
            logging.warning(
                "No 'Summary' sheet found in comparison summary %s; nothing copied.",
                comp_path,
            )
            return

        ws_comparison_sum = wb_comparison_sum["Summary"]

        # If Summary already exists in output, delete it
        if "Summary" in wb_output.sheetnames:
            del wb_output["Summary"]

        # Create Summary as the FIRST sheet
        ws_output = wb_output.create_sheet("Summary", 0)

        # Copy data + styles
        for row in ws_comparison_sum.iter_rows():
            for cell in row:
                new_cell = ws_output.cell(row=cell.row, column=cell.column, value=cell.value)

                # Copy individual style attributes (from your old script)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # Header colours for Bronze/Silver/Gold/Platinum
        header_colors = {
            "B1": "cd7f32",  # Bronze
            "C1": "C0C0C0",  # Silver
            "D1": "FFD700",  # Gold
            "E1": "E5E4E2",  # Platinum
        }

        for addr, color in header_colors.items():
            cell = ws_output[addr]
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.font = Font(bold=True, color="000000")

        wb_output.save(out_path)
        logging.debug(
            "Summary sheet copied to %s and placed as the first sheet with highlighted headers.",
            out_path,
        )

    except Exception as e:
        logging.error("Error in copy_summary_to_result: %s", e, exc_info=True)
        raise
