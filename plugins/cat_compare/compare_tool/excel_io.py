"""
excel_io.py
-----------
This module provides utility functions for working with Excel files.

Purpose:
- Handles operations such as saving workbooks and validating data consistency.
- Ensures formulas in Excel files are recalculated before processing.

Key Features:
- `save_workbook`: Opens and saves an Excel workbook using `xlwings` to ensure formulas are recalculated.
- `check_controllers_match`: Validates that two Excel files have matching controller values.
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook


def save_workbook(filepath: str) -> None:
    """
    Open and save the workbook in Excel so formulas are recalculated
    before we read it with pandas/openpyxl.
    """
    path = Path(filepath).resolve()
    logging.info("Saving workbook via Excel: %s", path)

    try:
        import xlwings as xw
    except ImportError as exc:
        raise RuntimeError(
            "xlwings is not installed, so Excel recalculation is unavailable."
        ) from exc

    app = xw.App(visible=False)
    wb = None
    try:
        wb = app.books.open(str(path))
        wb.save()
    finally:
        # Always try to close/quit even if something goes wrong
        try:
            wb.close()
        except Exception:
            pass
        app.quit()


def inspect_summary_formula_cache(filepath: str) -> Dict[str, object]:
    """
    Inspect whether Summary formula cells have cached values.

    openpyxl cannot calculate formulas. It can only read a formula's cached
    result if Excel or the file generator stored one in the workbook.
    """
    path = Path(filepath).resolve()
    result: Dict[str, object] = {
        "path": str(path),
        "summary_exists": False,
        "formula_cells": 0,
        "missing_cached_formula_cells": 0,
        "missing_cached_coordinates": [],
    }

    formula_wb = load_workbook(path, read_only=True, data_only=False)
    value_wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Summary" not in formula_wb.sheetnames:
            return result

        result["summary_exists"] = True
        formula_ws = formula_wb["Summary"]
        value_ws = value_wb["Summary"]
        missing: List[str] = []

        for row in formula_ws.iter_rows():
            for cell in row:
                value = cell.value
                is_formula = cell.data_type == "f" or (
                    isinstance(value, str) and value.startswith("=")
                )
                if not is_formula:
                    continue

                result["formula_cells"] = int(result["formula_cells"]) + 1
                cached_value = value_ws[cell.coordinate].value
                if cached_value is None:
                    missing.append(cell.coordinate)

        result["missing_cached_coordinates"] = missing
        result["missing_cached_formula_cells"] = len(missing)
        return result
    finally:
        formula_wb.close()
        value_wb.close()


def summary_cache_is_available(filepath: str) -> bool:
    """
    Return True when the Summary sheet either has no formulas or every Summary
    formula has a cached value available to openpyxl.
    """
    info = inspect_summary_formula_cache(filepath)
    return bool(info["summary_exists"]) and int(info["missing_cached_formula_cells"]) == 0


def summary_missing_cache_is_supported(filepath: str) -> bool:
    """
    Return True when every missing Summary cached value is one of the simple
    CAT Summary formulas we can calculate in Python.
    """
    info = inspect_summary_formula_cache(filepath)
    if not info["summary_exists"]:
        return False

    missing_coordinates = set(info["missing_cached_coordinates"])
    if not missing_coordinates:
        return True

    path = Path(filepath).resolve()
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        ws = wb["Summary"]
        for coordinate in missing_coordinates:
            formula = str(ws[coordinate].value or "").upper()
            if "COUNTIF(" in formula and "OVERALLASSESSMENT" in formula:
                continue
            if formula.startswith("=ROUND(") and "COUNTA(ANALYSIS!" in formula:
                continue
            return False
        return True
    finally:
        wb.close()


def check_controllers_match(previous_file_path: str, current_file_path: str) -> bool:
    """
    Ensure both workbooks have a single, matching controller value
    in the Analysis sheet's 'controller' column.
    """
    try:
        prev_df = pd.read_excel(previous_file_path, sheet_name="Analysis")
        curr_df = pd.read_excel(current_file_path, sheet_name="Analysis")
    except Exception as e:
        logging.error("Failed to read 'Analysis' sheet from one of the files: %s", e)
        return False

    if "controller" not in prev_df.columns or "controller" not in curr_df.columns:
        logging.error("Missing 'controller' column in one of the Analysis sheets.")
        return False

    prev_ctrls = prev_df["controller"].dropna().astype(str).str.strip().unique()
    curr_ctrls = curr_df["controller"].dropna().astype(str).str.strip().unique()

    logging.debug(f"Previous controller(s): {prev_ctrls}")
    logging.debug(f"Current controller(s): {curr_ctrls}")

    if len(prev_ctrls) != 1 or len(curr_ctrls) != 1:
        logging.error(
            "Controller column does not contain exactly one unique value in each file."
        )
        return False

    if prev_ctrls[0] != curr_ctrls[0]:
        logging.error(
            f"Controllers do not match: {prev_ctrls[0]} vs {curr_ctrls[0]}"
        )
        return False

    return True


def get_key_column(worksheet, header_name: str) -> Optional[int]:
    """
    Find the 1-based column index for a header in an openpyxl worksheet.
    Used by comparison routines to align columns.
    """
    try:
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1))
    except StopIteration:
        return None

    for cell in header_row:
        if str(cell.value or "").strip() == header_name:
            return cell.column
    return None
