"""
app.py
------
This is the main entry point for the Flask web application.

Purpose:
- Initializes the Flask app.
- Sets up routes for the homepage, insights, and file uploads.
- Configures logging and application folders.

Key Routes:
- `/`: Renders the homepage.
- `/insights`: Renders the insights page.
- `/upload`: Handles file uploads for APM comparisons.
"""

import os
import json
import re
import shutil
import threading
import webbrowser
import datetime as dt
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from flask import Flask, request, jsonify, render_template, send_from_directory
from openpyxl import load_workbook
from compare_tool.config import load_config
from compare_tool.logging_config import setup_logging
from compare_tool.insights import build_comparison_json
from compare_tool.service import (
    get_excel_recalculation_status,
    run_comparison,        # APM
    run_comparison_brum,   # BRUM
    run_comparison_mrum,   # MRUM
    find_best_matching_files,  # Folder processing
    save_matched_files,        # Folder processing
)
import logging


def ts_now():
    return dt.datetime.now(dt.timezone.utc).strftime("%Y%m%d_%H%M%S")


setup_logging()
logging.info("Logging setup is complete.")

BASE_DIR = Path(__file__).resolve().parent.parent  # points at compare-plugin
config = load_config(str(BASE_DIR / "config.json"))

UPLOAD_FOLDER = config["upload_folder"]   # e.g. "uploads"
RESULT_FOLDER = config["result_folder"]   # e.g. "results"
HISTORY_FOLDER = BASE_DIR / "history"     # used by insights APIs

app = Flask(
    __name__,
    static_folder=str(BASE_DIR / "static"),
    template_folder=str(BASE_DIR / "templates"),
)

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True)
os.makedirs(HISTORY_FOLDER, exist_ok=True)


def excel_status_note(domain: str) -> str:
    status = get_excel_recalculation_status(domain)
    if not status:
        return ""
    return f"<br><span style='color:#9fb3c8;'>{status}</span>"


def latest_summary_file(domain: str, controller: Optional[str] = None) -> Optional[str]:
    prefix = f"analysis_summary_{domain.lower()}_"
    if not os.path.isdir(RESULT_FOLDER):
        return None

    candidates = []
    for name in os.listdir(RESULT_FOLDER):
        if not (name.startswith(prefix) and name.endswith(".json")):
            continue
        if controller:
            path = os.path.join(RESULT_FOLDER, name)
            try:
                with open(path, "r", encoding="utf-8") as f:
                    payload = json.load(f)
                meta = payload.get("meta") or {}
            except Exception:
                continue
            if _slug(meta.get("controller")) != _slug(controller):
                continue
        candidates.append(name)

    if not candidates:
        return None
    return sorted(candidates)[-1]


def run_domain_comparison(domain: str, previous_path: str, current_path: str) -> Tuple[str, str]:
    domain = (domain or "").lower()
    if domain == "apm":
        return run_comparison(
            previous_file_path=previous_path,
            current_file_path=current_path,
            config=config,
        )
    if domain == "brum":
        return run_comparison_brum(
            previous_file_path=previous_path,
            current_file_path=current_path,
            config=config,
        )
    if domain == "mrum":
        return run_comparison_mrum(
            previous_file_path=previous_path,
            current_file_path=current_path,
            config=config,
        )
    raise ValueError(f"Unsupported domain: {domain}")


def _safe_slug(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", value or "").strip("_")
    return cleaned[:80] or "snapshot"


def copy_progression_output(path: str, domain: str, baseline: str, current: str, suffix: str) -> str:
    ext = Path(path).suffix
    name = (
        f"Progression_{domain.upper()}_"
        f"{_safe_slug(baseline)}_to_{_safe_slug(current)}_"
        f"{suffix}{ext}"
    )
    target = os.path.join(RESULT_FOLDER, name)
    shutil.copy2(path, target)
    return target


def relative_parts(filename: str) -> List[str]:
    return [p for p in re.split(r"[\\/]+", filename or "") if p]


MONTHS = {
    "jan": "01",
    "feb": "02",
    "mar": "03",
    "apr": "04",
    "may": "05",
    "jun": "06",
    "jul": "07",
    "aug": "08",
    "sep": "09",
    "sept": "09",
    "oct": "10",
    "nov": "11",
    "dec": "12",
}


def is_ignored_upload(filename: str) -> bool:
    parts = relative_parts(filename)
    basename = parts[-1] if parts else filename or ""
    lower = basename.lower()
    return (
        not basename
        or basename.startswith(".")
        or basename.startswith("~$")
        or lower in {"thumbs.db", "desktop.ini"}
    )


def detect_assessment_date(value: str) -> Optional[str]:
    match = re.search(r"(20\d{2})[-_ .]?([01]\d)[-_ .]?([0-3]\d)", value or "")
    if match:
        return match.group(1) + match.group(2) + match.group(3)
    match = re.search(r"\b(20\d{6})\b", value or "")
    if match:
        return match.group(1)
    match = re.search(r"\b([01]?\d)[-_ .]([0-3]?\d)[-_ .](\d{2})\b", value or "")
    if match:
        month = int(match.group(1))
        day = int(match.group(2))
        year = 2000 + int(match.group(3))
        if 1 <= month <= 12 and 1 <= day <= 31:
            return f"{year:04d}{month:02d}{day:02d}"
    match = re.search(
        r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*[-_ .]?(\d{2,4})\b",
        value or "",
        re.IGNORECASE,
    )
    if match:
        month = MONTHS[match.group(1).lower()[:3]]
        year_text = match.group(2)
        year = int(year_text) if len(year_text) == 4 else 2000 + int(year_text)
        return f"{year:04d}{month}01"
    return None


def format_workbook_date(value: object) -> Optional[str]:
    if isinstance(value, dt.datetime):
        return value.strftime("%Y%m%d")
    if isinstance(value, dt.date):
        return value.strftime("%Y%m%d")
    if value:
        return detect_assessment_date(str(value))
    return None


def sheet_by_name(workbook, wanted: str):
    wanted_lower = wanted.strip().lower()
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip().lower() == wanted_lower:
            return workbook[sheet_name]
    return None


def cell_text(value: object) -> str:
    return str(value or "").strip()


def normalized_cell_text(value: object) -> str:
    return cell_text(value).lstrip("\ufeff").lower()


def scan_sheet_date(workbook) -> Optional[str]:
    date_labels = ("date", "generated", "created", "export", "assessment")
    for sheet_name in workbook.sheetnames[:6]:
        ws = workbook[sheet_name]
        max_row = min(ws.max_row or 1, 30)
        max_col = min(ws.max_column or 1, 12)
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
            values = list(row)
            for idx, value in enumerate(values):
                direct = format_workbook_date(value)
                if direct and any(label in cell_text(other).lower() for other in values for label in date_labels):
                    return direct
                text = cell_text(value).lower()
                if not any(label in text for label in date_labels):
                    continue
                for nearby in values[idx:idx + 4]:
                    found = format_workbook_date(nearby)
                    if found:
                        return found
    return None


def scan_workbook_controller(workbook) -> Optional[str]:
    ws = sheet_by_name(workbook, "Analysis")
    if ws is None:
        return None

    if normalized_cell_text(ws["A1"].value) == "controller":
        controllers = []
        for row in ws.iter_rows(
            min_row=2,
            max_row=min(ws.max_row or 2, 5001),
            min_col=1,
            max_col=1,
            values_only=True,
        ):
            value = cell_text(row[0] if row else "")
            if value and normalized_cell_text(value) != "controller" and value not in controllers:
                controllers.append(value)
            if len(controllers) > 1:
                return "Multiple controllers"
        return controllers[0] if controllers else None

    header_row = None
    controller_col = None
    max_header_row = min(ws.max_row or 1, 20)
    for row in ws.iter_rows(min_row=1, max_row=max_header_row):
        for cell in row:
            if normalized_cell_text(cell.value) == "controller":
                header_row = cell.row
                controller_col = cell.column
                break
        if controller_col:
            break

    if not header_row or not controller_col:
        return None

    controllers = []
    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_row=min(ws.max_row or header_row + 1, header_row + 5000),
        min_col=controller_col,
        max_col=controller_col,
        values_only=True,
    ):
        value = cell_text(row[0] if row else "")
        if value and normalized_cell_text(value) != "controller" and value not in controllers:
            controllers.append(value)
        if len(controllers) > 1:
            return "Multiple controllers"
    return controllers[0] if controllers else None


def workbook_preview_metadata(file_obj: object) -> Dict[str, object]:
    cached = getattr(file_obj, "_cat_preview_metadata", None)
    if cached is not None:
        return cached

    stream = getattr(file_obj, "stream", None)
    if stream is None:
        return {"date": None, "controller": None}

    metadata: Dict[str, object] = {
        "date": None,
        "controller": None,
    }
    try:
        stream.seek(0)
        workbook_data = stream.read()
        workbook = load_workbook(BytesIO(workbook_data), read_only=True, data_only=True)
        metadata["date"] = (
            format_workbook_date(workbook.properties.created)
            or format_workbook_date(workbook.properties.modified)
            or scan_sheet_date(workbook)
        )
        metadata["controller"] = scan_workbook_controller(workbook)
        workbook.close()
    except Exception:
        pass
    finally:
        try:
            stream.seek(0)
        except Exception:
            pass

    try:
        setattr(file_obj, "_cat_preview_metadata", metadata)
    except Exception:
        pass
    return metadata


def workbook_metadata_date(file_obj: object) -> Optional[str]:
    return workbook_preview_metadata(file_obj).get("date")


def workbook_controller(file_obj: object) -> Optional[str]:
    return workbook_preview_metadata(file_obj).get("controller")


def workbook_metadata_assessment_date(files: List[object]) -> Optional[str]:
    dated_files = []
    for file in files:
        if not domain_for_filename(getattr(file, "filename", "")):
            continue
        file_date = workbook_metadata_date(file)
        if file_date:
            dated_files.append(file_date)
    return min(dated_files) if dated_files else None


def best_assessment_date(label: str, files: List[object]) -> Optional[str]:
    metadata_date = workbook_metadata_assessment_date(files)
    if metadata_date:
        return metadata_date
    date = detect_assessment_date(label)
    if date:
        return date
    dated_files = []
    for file in files:
        filename = getattr(file, "filename", "")
        file_date = detect_assessment_date(filename) or workbook_metadata_date(file)
        if file_date:
            dated_files.append(file_date)
    return min(dated_files) if dated_files else None


def domain_for_filename(filename: str) -> Optional[str]:
    text = (filename or "").lower()
    if "raw" in text or "maturityassessment" not in text:
        return None
    if "apm" in text:
        return "APM"
    if "brum" in text:
        return "BRUM"
    if "mrum" in text:
        return "MRUM"
    return None


def uploaded_file_modified_date(filename: str, uploaded_dates: Optional[Dict[str, str]]) -> Optional[str]:
    if not uploaded_dates:
        return None
    exact = uploaded_dates.get(filename)
    if exact:
        return exact
    normalized = "/".join(relative_parts(filename))
    return uploaded_dates.get(normalized)


def progression_group_summary(name: str, files: List[object], uploaded_dates: Optional[Dict[str, str]] = None) -> dict:
    domain_files = [
        f for f in files
        if domain_for_filename(getattr(f, "filename", ""))
    ]
    metadata_date = workbook_metadata_assessment_date(domain_files)
    modified_dates = sorted(
        {
            date
            for date in (
                uploaded_file_modified_date(getattr(f, "filename", ""), uploaded_dates)
                for f in domain_files
            )
            if date
        }
    )
    modified_date = modified_dates[0] if modified_dates else None
    explicit_date = None if metadata_date or modified_date else detect_assessment_date(" ".join([name] + [getattr(f, "filename", "") for f in files]))
    domains = sorted(
        {
            domain
            for domain in (domain_for_filename(getattr(f, "filename", "")) for f in domain_files)
            if domain
        }
    )
    controllers = sorted(
        {
            controller
            for controller in (workbook_controller(f) for f in domain_files)
            if controller
        }
    )
    controller_hints = sorted(
        {
            hint
            for hint in (detect_controller_hint(getattr(f, "filename", "")) for f in files)
            if hint
        }
    )[:4]
    controller = (
        controllers[0]
        if len(controllers) == 1
        else "Mixed"
        if len(controllers) > 1
        else "Not inspected"
        if not domain_files
        else "Unknown"
    )
    nested_folder = any(
        len(relative_parts(getattr(file, "filename", ""))) > 3
        for file in files
    )
    return {
        "name": name,
        "date": metadata_date or modified_date or explicit_date or "Unknown",
        "dateSource": (
            "workbook metadata"
            if metadata_date
            else "file modified"
            if modified_date
            else "folder/filename"
            if explicit_date
            else "unknown"
        ),
        "domains": domains,
        "controllers": controllers,
        "controller": controller,
        "controllerHints": controller_hints,
        "nestedFolder": nested_folder,
        "nestedFolderMessage": (
            "Nested folder detected. Browse one level deeper if this row is not the assessment set you want."
            if nested_folder
            else ""
        ),
    }


def detect_controller_hint(value: str) -> str:
    parts = relative_parts(value)
    basename = Path(parts[-1] if parts else value or "").stem
    tokens = re.split(r"[^A-Za-z0-9.]+", basename)
    stop = {
        "apm",
        "brum",
        "mrum",
        "raw",
        "maturityassessment",
        "maturity",
        "assessment",
        "controller",
        "config",
        "configuration",
        "export",
        "report",
    }
    cleaned = []
    for token in tokens:
        low = token.lower()
        if not token or low in stop:
            continue
        if detect_assessment_date(token):
            continue
        cleaned.append(token)
    if not cleaned:
        return ""
    return "_".join(cleaned[:3])


def group_uploaded_assessment_folders(files: List[object]) -> List[Tuple[str, List[object]]]:
    valid = [
        f for f in files
        if getattr(f, "filename", "") and not is_ignored_upload(getattr(f, "filename", ""))
    ]
    if not valid:
        return []

    def flat_group_name(file: object) -> Optional[str]:
        filename = getattr(file, "filename", "")
        if not domain_for_filename(filename):
            return None
        date = workbook_metadata_date(file) or detect_assessment_date(filename)
        if not date:
            return None
        controller = workbook_controller(file)
        controller_hint = controller or detect_controller_hint(filename)
        return f"{controller_hint} {date}".strip() or date

    def add_flat_groups(grouped: dict, flat_files: List[object]) -> None:
        for file in flat_files:
            group_name = flat_group_name(file)
            if not group_name:
                continue
            grouped.setdefault(group_name, []).append(file)

    split_names = [relative_parts(getattr(f, "filename", "")) for f in valid]
    first_segments = [parts[0] for parts in split_names if parts]
    has_common_root = bool(first_segments) and len(set(first_segments)) == 1

    if has_common_root and any(len(parts) > 2 for parts in split_names):
        grouped = {}
        loose_files = []
        for file in valid:
            parts = relative_parts(getattr(file, "filename", ""))
            if len(parts) < 3:
                loose_files.append(file)
                continue
            group_name = parts[1]
            grouped.setdefault(group_name, []).append(file)
        add_flat_groups(grouped, loose_files)
        if len(grouped) >= 2:
            return sorted(grouped.items(), key=lambda item: folder_sort_key(item[0], item[1]))

    grouped = {}
    add_flat_groups(grouped, valid)

    return sorted(grouped.items(), key=lambda item: folder_sort_key(item[0], item[1]))


def folder_sort_key(folder_name: str, files: List[object]):
    return (best_assessment_date(folder_name, files) or "99999999", folder_name.lower())


def has_nested_assessment_folders(files: List[object]) -> bool:
    valid_parts = [
        relative_parts(getattr(file, "filename", ""))
        for file in files
        if getattr(file, "filename", "") and not is_ignored_upload(getattr(file, "filename", ""))
    ]
    roots = [parts[0] for parts in valid_parts if parts]
    has_common_root = bool(roots) and len(set(roots)) == 1
    if not has_common_root:
        return False
    return any(len(parts) > 3 for parts in valid_parts)


def progression_preview_sort_key(group: dict):
    controller = group.get("controller") or ""
    controller_key = controller.lower()
    if controller in {"Unknown", "Mixed", "Not inspected"}:
        controller_key = f"zz_{controller_key}"
    date_key = group.get("date") or "99999999"
    if date_key == "Unknown":
        date_key = "99999999"
    return (controller_key, date_key, (group.get("name") or "").lower())


@app.route("/", methods=["GET"])
def index():
    return render_template("index.html", active_tab="apm")


@app.route("/insights", methods=["GET"])
def insights():
    return render_template("insights.html")


@app.route("/portfolio", methods=["GET"])
def portfolio():
    return render_template("portfolio.html")


# ---------- APM upload (uses new compare_tool.service) -----------------------
@app.route("/upload", methods=["POST"])
def upload_apm():
    if "previous_file" not in request.files or "current_file" not in request.files:
        return render_template("index.html", message="Missing files.", active_tab="apm"), 400

    prev = request.files["previous_file"]
    curr = request.files["current_file"]

    if not prev.filename or not curr.filename:
        return render_template("index.html", message="Please select both files.", active_tab="apm"), 400

    prev_path = os.path.join(UPLOAD_FOLDER, "previous_apm.xlsx")
    curr_path = os.path.join(UPLOAD_FOLDER, "current_apm.xlsx")

    prev.save(prev_path)
    curr.save(curr_path)

    # Run the APM comparison pipeline
    output_file, ppt_file = run_comparison(
        previous_file_path=prev_path,
        current_file_path=curr_path,
        config=config,
    )

    # 🔹 Build JSON snapshot in RESULT_FOLDER so /api/history can see it
    json_path, json_name, _ = build_comparison_json(
        domain="APM",
        comparison_result_path=output_file,
        current_file_path=curr_path,
        previous_file_path=prev_path,
        result_folder=RESULT_FOLDER,   # IMPORTANT
    )

    app.config["LAST_RESULT_APM"] = output_file
    app.config["LAST_PPT_APM"] = ppt_file
    app.config["LAST_JSON_APM"] = json_path

    msg = (
        "APM comparison completed. "
        f"Download Excel <a href='/download/{os.path.basename(output_file)}' style='color:#32CD32;'>here</a> "
        f"and PowerPoint <a href='/download/{os.path.basename(ppt_file)}' style='color:#32CD32;'>here</a>. "
        "Insights snapshot has been generated and will be available on the Insights page."
        f"{excel_status_note('APM')}"
    )
    return render_template("index.html", message=msg, active_tab="apm")




@app.route("/download/<filename>")
def download(filename):
    return send_from_directory(RESULT_FOLDER, filename, as_attachment=True)


# ---------- BRUM / MRUM upload placeholders ---------------------------------
@app.route("/upload_brum", methods=["POST"])
def upload_brum():
    if "previous_brum" not in request.files or "current_brum" not in request.files:
        return render_template("index.html", message="Missing BRUM files.", active_tab="brum"), 400

    prev = request.files["previous_brum"]
    curr = request.files["current_brum"]

    if not prev.filename or not curr.filename:
        return render_template("index.html", message="Please select both BRUM files.", active_tab="brum"), 400

    prev_path = os.path.join(UPLOAD_FOLDER, "previous_brum.xlsx")
    curr_path = os.path.join(UPLOAD_FOLDER, "current_brum.xlsx")

    prev.save(prev_path)
    curr.save(curr_path)

    output_file, ppt_file = run_comparison_brum(
        previous_file_path=prev_path,
        current_file_path=curr_path,
        config=config,
    )

    json_path, json_name, _ = build_comparison_json(
        domain="BRUM",
        comparison_result_path=output_file,
        current_file_path=curr_path,
        previous_file_path=prev_path,
        result_folder=RESULT_FOLDER,
    )

    msg = (
        "BRUM comparison completed. "
        f"Download Excel <a href='/download/{os.path.basename(output_file)}' style='color:#32CD32;'>here</a> "
        f"and PowerPoint <a href='/download/{os.path.basename(ppt_file)}' style='color:#32CD32;'>here</a>. "
        "BRUM Insights snapshot has been generated and will be available on the Insights page."
        f"{excel_status_note('BRUM')}"
    )
    return render_template("index.html", message=msg, active_tab="brum")


@app.route("/upload_mrum", methods=["POST"])
def upload_mrum():
    if "previous_mrum" not in request.files or "current_mrum" not in request.files:
        return render_template("index.html", message="Missing MRUM files.", active_tab="mrum"), 400

    prev = request.files["previous_mrum"]
    curr = request.files["current_mrum"]

    if not prev.filename or not curr.filename:
        return render_template("index.html", message="Please select both MRUM files.", active_tab="mrum"), 400

    prev_path = os.path.join(UPLOAD_FOLDER, "previous_mrum.xlsx")
    curr_path = os.path.join(UPLOAD_FOLDER, "current_mrum.xlsx")

    prev.save(prev_path)
    curr.save(curr_path)

    output_file, ppt_file = run_comparison_mrum(
        previous_file_path=prev_path,
        current_file_path=curr_path,
        config=config,
    )

    json_path, json_name, _ = build_comparison_json(
        domain="MRUM",
        comparison_result_path=output_file,
        current_file_path=curr_path,
        previous_file_path=prev_path,
        result_folder=RESULT_FOLDER,
    )

    msg = (
        "MRUM comparison completed. "
        f"Download Excel <a href='/download/{os.path.basename(output_file)}' style='color:#32CD32;'>here</a> "
        f"and PowerPoint <a href='/download/{os.path.basename(ppt_file)}' style='color:#32CD32;'>here</a>. "
        "MRUM Insights snapshot has been generated and will be available on the Insights page."
        f"{excel_status_note('MRUM')}"
    )
    return render_template("index.html", message=msg, active_tab="mrum")


# ---------- Folder upload (processes multiple data types) --------------------
@app.route("/upload_folders", methods=["POST"])
def upload_folders():
    logging.debug("[FOLDERS] Request files: %s", list(request.files.keys()))
    
    # Check if folders were uploaded
    if 'previous_folder' not in request.files or 'current_folder' not in request.files:
        logging.error("[FOLDERS] No folder part")
        return render_template('index.html', message="Error: Please select both previous and current folders.", active_tab="folders"), 400
    
    # Get the selected data types from checkboxes
    selected_types = request.form.getlist('data_types')
    if not selected_types:
        return render_template('index.html', message="Error: Please select at least one data type (APM, BRUM, or MRUM).", active_tab="folders"), 400
    
    logging.info(f"[FOLDERS] Selected data types: {selected_types}")
    
    # Get all files from both folders
    previous_files = request.files.getlist('previous_folder')
    current_files = request.files.getlist('current_folder')
    
    logging.info(f"[FOLDERS] Previous folder: {len(previous_files)} files")
    logging.info(f"[FOLDERS] Current folder: {len(current_files)} files")
    
    # Find matching files for each data type
    matches = find_best_matching_files(previous_files, current_files)
    
    # Process each selected data type
    results = {}
    errors = []
    
    for data_type in selected_types:
        domain = data_type.upper()
        logging.info(f"[FOLDERS] Processing {domain}")
        
        try:
            # Save matched files for this domain
            previous_path, current_path = save_matched_files(matches, UPLOAD_FOLDER, data_type)
            
            if not previous_path or not current_path:
                errors.append(f"No matching {domain} files found in the selected folders.")
                continue
            
            output_file, ppt_file = run_domain_comparison(data_type, previous_path, current_path)
            
            # Build JSON snapshot for insights
            json_path, json_name, _ = build_comparison_json(
                domain=domain,
                comparison_result_path=output_file,
                current_file_path=current_path,
                previous_file_path=previous_path,
                result_folder=RESULT_FOLDER,
            )
            
            # Store results
            results[domain] = {
                'xlsx': os.path.basename(output_file),
                'pptx': os.path.basename(ppt_file),
                'json': json_name,
                'excel_status': get_excel_recalculation_status(domain)
            }
            
            logging.info(f"[FOLDERS] Successfully processed {domain}")
            
        except Exception as e:
            logging.error(f"[FOLDERS] Error processing {domain}: {e}", exc_info=True)
            errors.append(f"{domain}: Error during processing - {str(e)}")
    
    # Generate response message
    if results:
        message_parts = ["Processing completed successfully!<br><br>"]
        
        for domain, files in results.items():
            message_parts.append(f"<strong>{domain}:</strong><br>")
            message_parts.append(f"• Results: <a href='/download/{files['xlsx']}' style='color: #32CD32;'>Download Excel</a><br>")
            message_parts.append(f"• PowerPoint: <a href='/download/{files['pptx']}' style='color: #32CD32;'>Download PPT</a><br>")
            message_parts.append(f"• JSON: <a href='/download/{files['json']}' style='color: #32CD32;'>Download JSON</a><br><br>")
            if files.get('excel_status'):
                message_parts.append(f"<span style='color:#9fb3c8;'>{files['excel_status']}</span><br><br>")
        
        if errors:
            message_parts.append("<br><strong>Warnings:</strong><br>")
            for error in errors:
                message_parts.append(f"• {error}<br>")
        
        message = "".join(message_parts)
    else:
        message = f"Error: No files could be processed. Issues encountered:<br>{'<br>'.join(errors)}"
        return render_template('index.html', message=message, active_tab="folders"), 400
    
    return render_template('index.html', message=message, active_tab="folders")


@app.route("/upload_progression_folders", methods=["POST"])
def upload_progression_folders():
    if "progression_folder" not in request.files:
        return render_template(
            "index.html",
            message="Error: Please select a parent folder containing multiple assessment folders.",
            active_tab="folders",
        ), 400

    selected_types = request.form.getlist("progression_data_types")
    if not selected_types:
        return render_template(
            "index.html",
            message="Error: Please select at least one data type for progression compare.",
            active_tab="folders",
        ), 400

    grouped = group_uploaded_assessment_folders(request.files.getlist("progression_folder"))
    baseline_group = request.form.get("progression_baseline_group", "").strip()
    selected_groups_raw = request.form.get("progression_selected_groups", "")
    selected_groups = {
        name for name in selected_groups_raw.split("||")
        if name.strip()
    }
    grouped_by_name = {name: files for name, files in grouped}

    if not baseline_group or baseline_group not in grouped_by_name:
        return render_template(
            "index.html",
            message="Error: Select the baseline group before running progression compare.",
            active_tab="folders",
        ), 400

    selected_groups.discard(baseline_group)
    compare_groups = [(name, files) for name, files in grouped if name in selected_groups]

    if len(compare_groups) < 1:
        return render_template(
            "index.html",
            message=(
                "Error: Progression Compare needs a selected baseline and at least one selected comparison group."
            ),
            active_tab="folders",
        ), 400

    baseline_name, baseline_files = baseline_group, grouped_by_name[baseline_group]
    run_suffix_base = ts_now()
    results = []
    errors = []

    logging.info(
        "[PROGRESSION] Baseline folder=%s, current folders=%s, selected=%s",
        baseline_name,
        [name for name, _ in compare_groups],
        selected_types,
    )

    for current_index, (current_name, current_files) in enumerate(compare_groups, start=1):
        matches = find_best_matching_files(baseline_files, current_files)
        for data_type in selected_types:
            domain = data_type.upper()
            try:
                previous_path, current_path = save_matched_files(matches, UPLOAD_FOLDER, data_type)
                if not previous_path or not current_path:
                    errors.append(
                        f"{domain}: No matching files found for {baseline_name} to {current_name}."
                    )
                    continue

                output_file, ppt_file = run_domain_comparison(data_type, previous_path, current_path)
                pair_suffix = f"{run_suffix_base}_{current_index:02d}_{domain.lower()}"
                stable_output = copy_progression_output(
                    output_file, domain, baseline_name, current_name, pair_suffix
                )
                stable_ppt = copy_progression_output(
                    ppt_file, domain, baseline_name, current_name, pair_suffix
                )

                json_path, json_name, _ = build_comparison_json(
                    domain=domain,
                    comparison_result_path=output_file,
                    current_file_path=current_path,
                    previous_file_path=previous_path,
                    result_folder=RESULT_FOLDER,
                    meta={"compareDate": f"{run_suffix_base}_{current_index:02d}_{domain.lower()}"},
                )

                results.append(
                    {
                        "domain": domain,
                        "baseline": baseline_name,
                        "current": current_name,
                        "xlsx": os.path.basename(stable_output),
                        "pptx": os.path.basename(stable_ppt),
                        "json": json_name,
                        "excel_status": get_excel_recalculation_status(domain),
                    }
                )
            except Exception as e:
                logging.error(
                    "[PROGRESSION] Error processing %s %s -> %s: %s",
                    domain,
                    baseline_name,
                    current_name,
                    e,
                    exc_info=True,
                )
                errors.append(f"{domain} {baseline_name} to {current_name}: {str(e)}")

    if results:
        parts = [
            "<strong>Progression Compare completed.</strong><br>",
            f"Baseline group: <strong>{baseline_name}</strong><br>",
            f"Compared against {len(compare_groups)} selected group(s).<br><br>",
        ]
        for row in results:
            parts.append(
                f"<strong>{row['domain']}:</strong> "
                f"{row['baseline']} → {row['current']}<br>"
            )
            parts.append(
                f"• Results: <a href='/download/{row['xlsx']}' style='color:#32CD32;'>Excel</a> "
                f"• PowerPoint: <a href='/download/{row['pptx']}' style='color:#32CD32;'>PPT</a> "
                f"• JSON: <a href='/download/{row['json']}' style='color:#32CD32;'>JSON</a><br>"
            )
            if row.get("excel_status"):
                parts.append(f"<span style='color:#9fb3c8;'>{row['excel_status']}</span><br>")
            parts.append("<br>")
        if errors:
            parts.append("<strong>Warnings:</strong><br>")
            for error in errors:
                parts.append(f"• {error}<br>")
        message = "".join(parts)
        return render_template("index.html", message=message, active_tab="folders")

    message = "Error: No progression comparisons could be processed.<br>" + "<br>".join(errors)
    return render_template("index.html", message=message, active_tab="folders"), 400


@app.route("/api/progression_preview", methods=["POST"])
def api_progression_preview():
    if "progression_folder" not in request.files:
        return jsonify({"groups": [], "warnings": ["No folder files were provided."]}), 400

    files = request.files.getlist("progression_folder")
    uploaded_dates = {}
    for entry in request.form.getlist("progression_file_modified"):
        if "||" not in entry:
            continue
        filename, value = entry.split("||", 1)
        if value:
            uploaded_dates[filename] = value

    grouped = group_uploaded_assessment_folders(files)
    groups = []
    warnings = []
    if has_nested_assessment_folders(files):
        warnings.append(
            "Nested folders detected. For best results, browse one level deeper and select the folder containing the assessment folders or workbooks."
        )

    for index, (name, group_files) in enumerate(grouped):
        summary = progression_group_summary(name, group_files, uploaded_dates)
        summary["role"] = "Select"
        summary["includeDefault"] = bool(summary["domains"])
        summary["baselineDefault"] = False
        if not summary["domains"]:
            summary["reason"] = "Excluded: no APM, BRUM, or MRUM MaturityAssessment files found."
        else:
            summary["reason"] = "Select as baseline if this is the earliest assessment, or compare against the chosen baseline."
        groups.append(summary)

    groups.sort(key=progression_preview_sort_key)

    if len(groups) < 2:
        warnings.append(
            "Could not find at least two assessment groups. Use subfolders, filename dates, or workbook metadata."
        )
    inspectable_groups = [g for g in groups if g.get("domains")]
    controller_values = {
        g["controller"]
        for g in inspectable_groups
        if g.get("controller") not in {"Unknown", "Mixed", "Not inspected"}
    }
    if len(controller_values) > 1:
        warnings.append(
            "Different workbook controllers were found. Untick groups that do not belong to the baseline controller."
        )
    if any(g.get("controller") in {"Unknown", "Mixed"} for g in inspectable_groups):
        warnings.append(
            "Some controllers could not be read or were mixed. The final compare will still block mismatched controllers."
        )

    return jsonify({"groups": groups, "warnings": warnings})


#####################################################################################
############## Utility for Index on Read (compare multiple output) ##################
#####################################################################################


def _slug(s: Optional[str]) -> str:
    if not s:
        return ""
    return "".join(ch.lower() for ch in s if ch.isalnum())


def scan_runs(folder: str, domain: str, controller_filter: Optional[str], limit: int):
    """
    Scan RESULT_FOLDER for analysis_summary_<domain>_*.json
    and build a list of run dicts for trends.
    """
    prefix = f"analysis_summary_{domain}_"
    if not os.path.isdir(folder):
        return []

    runs = []
    for name in sorted(os.listdir(folder), reverse=True):
        if not (name.startswith(prefix) and name.endswith(".json")):
            continue

        path = os.path.join(folder, name)
        try:
            with open(path, "r", encoding="utf-8") as f:
                payload = json.load(f)
            meta = payload.get("meta") or {}
        except Exception:
            continue

        controller = meta.get("controller")
        if controller_filter and _slug(controller) != _slug(controller_filter):
            continue

        previousDate = meta.get("previousDate") or ""
        currentDate = meta.get("currentDate") or ""
        compareDate = meta.get("compareDate") or ""

        improved = int(meta.get("improved", 0))
        degraded = int(meta.get("degraded", 0))
        percentage = float(meta.get("percentage", 0.0))
        tiers = meta.get("tiers") or {}

        runs.append(
            {
                "file": name,
                "controller": controller,
                "previousDate": previousDate,
                "currentDate": currentDate,
                "compareDate": compareDate,
                "improved": improved,
                "degraded": degraded,
                "percentage": percentage,
                "tiers": tiers,
                "sortPrev": previousDate,
            }
        )

    # newest compareDate first
    runs.sort(key=lambda r: r["compareDate"], reverse=True)
    return runs[:limit]


def _run_sort_key(item):
    return (
        item.get("currentDate") or "",
        item.get("compareDate") or "",
        item.get("previousDate") or "",
    )


def fixed_baseline_runs(runs):
    """
    Keep only trend points that start from the earliest previousDate, de-dupe
    repeated previous/current pairs, and order by currentDate.
    """
    prev_dates = [r.get("previousDate") for r in runs if r.get("previousDate")]
    if not prev_dates:
        return sorted(runs, key=_run_sort_key)

    baseline = min(prev_dates)
    deduped = {}
    for run in runs:
        if run.get("previousDate") != baseline:
            continue
        key = (run.get("previousDate") or "", run.get("currentDate") or "")
        existing = deduped.get(key)
        if not existing or (run.get("compareDate") or "") > (existing.get("compareDate") or ""):
            deduped[key] = run

    return sorted(deduped.values(), key=_run_sort_key)


# ---------- Insights API stubs (match your JS expectations) ------------------
# These should read/write JSON files under HISTORY_FOLDER.
# For now, you can leave your existing implementations here and just
# update them later to use the new comparison outputs.

@app.route("/api/history", methods=["GET"])
def api_history():
    """
    Return a list of available JSON snapshots for the given domain.

    Looks in RESULT_FOLDER for files like:
      analysis_summary_<domain>_YYYYMMDD_HHMMSS.json
    and exposes light metadata used by the Insights UI.
    """
    domain = (request.args.get("domain") or "").lower()
    if domain not in ("apm", "brum", "mrum"):
        return jsonify({"error": "Invalid domain."}), 400
    include_duplicates = (request.args.get("include_duplicates") or "").lower() in ("1", "true", "yes")

    folder = RESULT_FOLDER
    prefix = f"analysis_summary_{domain}_"   # <-- matches your filenames

    items = []

    if not os.path.isdir(folder):
        return jsonify({"domain": domain.upper(), "items": []})

    for name in sorted(os.listdir(folder), reverse=True):
        if not (name.startswith(prefix) and name.endswith(".json")):
            continue

        path = os.path.join(folder, name)
        meta = {}
        try:
            with open(path, "r", encoding="utf-8") as f:
                payload = json.load(f)
            meta = payload.get("meta") or {}
        except Exception:
            meta = {}

        items.append(
            {
                "file": name,
                "timestamp": meta.get("compareDate", ""),
                "controller": meta.get("controller"),
                "prev": meta.get("previousDate"),
                "curr": meta.get("currentDate"),
            }
        )

    items.sort(key=lambda x: x["timestamp"] or "", reverse=True)

    # optional controller filter
    controller_q = request.args.get("controller")
    if controller_q:
        want = _slug(controller_q)
        filtered = []
        for it in items:
            if it["controller"] and _slug(it["controller"]) == want:
                filtered.append(it)
        items = filtered

    duplicate_counts = {}
    duplicate_files = {}
    for item in items:
        key = (
            _slug(item.get("controller")),
            item.get("prev") or "",
            item.get("curr") or "",
        )
        duplicate_counts[key] = duplicate_counts.get(key, 0) + 1
        duplicate_files.setdefault(key, []).append(item["file"])

    for item in items:
        key = (
            _slug(item.get("controller")),
            item.get("prev") or "",
            item.get("curr") or "",
        )
        item["duplicateCount"] = duplicate_counts.get(key, 1)
        item["duplicateFiles"] = duplicate_files.get(key, [item["file"]])

    if not include_duplicates:
        deduped = []
        seen = set()
        for item in items:
            key = (
                _slug(item.get("controller")),
                item.get("prev") or "",
                item.get("curr") or "",
            )
            if key in seen:
                continue
            seen.add(key)
            deduped.append(item)
        items = deduped

    controllers = sorted(
        {it["controller"] for it in items if it.get("controller")},
        key=lambda value: value.lower(),
    )
    hidden_duplicates = sum(max(0, count - 1) for count in duplicate_counts.values())
    return jsonify(
        {
            "domain": domain.upper(),
            "items": items,
            "controllers": controllers,
            "hiddenDuplicates": 0 if include_duplicates else hidden_duplicates,
        }
    )


@app.route("/api/apps", methods=["GET"])
def api_apps():
    """
    Return list of application names for a given domain & snapshot.

    If ?file=<name> is not provided, uses the latest snapshot for that domain.
    """
    domain = (request.args.get("domain") or "APM").upper()
    folder = RESULT_FOLDER

    # Optional explicit file selection
    file_name = request.args.get("file")
    controller = request.args.get("controller")
    include_status = (request.args.get("include_status") or "").lower() in ("1", "true", "yes")

    def _latest_file_for_domain() -> Optional[str]:
        return latest_summary_file(domain, controller)

    if not file_name:
        file_name = _latest_file_for_domain()

    if not file_name:
        # No snapshots yet for this domain
        return jsonify({"apps": []})

    path = os.path.join(folder, file_name)
    if not os.path.exists(path):
        return jsonify({"apps": []})

    try:
        with open(path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        apps = payload.get("apps", {}).get("names", []) or []
        if include_status:
            apps_index = payload.get("appsIndex") or {}
            enriched = []
            for app_name in apps:
                entry = apps_index.get(app_name) or {}
                upgraded = 0
                downgraded = 0
                unchanged = 0
                for area in entry.get("areas") or []:
                    status = str(area.get("status") or "").lower()
                    if status == "upgraded":
                        upgraded += 1
                    elif status == "downgraded":
                        downgraded += 1
                    else:
                        unchanged += 1
                enriched.append(
                    {
                        "name": app_name,
                        "upgraded": upgraded,
                        "downgraded": downgraded,
                        "unchanged": unchanged,
                        "changed": upgraded + downgraded,
                        "net": upgraded - downgraded,
                    }
                )
            apps = enriched
    except Exception:
        apps = []

    return jsonify({"apps": apps})


@app.route("/api/portfolio", methods=["GET"])
def api_portfolio():
    domain = (request.args.get("domain") or "APM").upper()
    if domain not in ("APM", "BRUM", "MRUM"):
        return jsonify({"error": "Invalid domain."}), 400

    file_name = request.args.get("file") or ""
    controller = request.args.get("controller") or ""
    if not file_name:
        file_name = latest_summary_file(domain, controller or None) or ""

    if not file_name:
        return jsonify({"domain": domain, "apps": [], "meta": {}})

    path = os.path.join(RESULT_FOLDER, file_name)
    if not os.path.exists(path):
        return jsonify({"domain": domain, "apps": [], "meta": {}})

    try:
        with open(path, "r", encoding="utf-8") as f:
            payload = json.load(f)
    except Exception:
        return jsonify({"domain": domain, "apps": [], "meta": {}})

    apps_index = payload.get("appsIndex") or {}
    app_names = payload.get("apps", {}).get("names", []) or sorted(apps_index.keys())
    apps = []

    for app_name in app_names:
        entry = apps_index.get(app_name) or {}
        upgraded = 0
        downgraded = 0
        unchanged = 0
        for area in entry.get("areas") or []:
            status = str(area.get("status") or "").lower()
            if status == "upgraded":
                upgraded += 1
            elif status == "downgraded":
                downgraded += 1
            else:
                unchanged += 1

        if downgraded and upgraded:
            status = "Mixed"
            severity = 2
        elif downgraded:
            status = "Degraded"
            severity = 3
        elif upgraded:
            status = "Improved"
            severity = 1
        else:
            status = "No Change"
            severity = 0

        apps.append(
            {
                "name": app_name,
                "status": status,
                "severity": severity,
                "upgraded": upgraded,
                "downgraded": downgraded,
                "unchanged": unchanged,
                "changed": upgraded + downgraded,
            }
        )

    apps.sort(key=lambda item: (-item["severity"], -item["changed"], item["name"].lower()))
    return jsonify(
        {
            "domain": domain,
            "file": file_name,
            "meta": payload.get("meta", {}),
            "apps": apps,
        }
    )


@app.route("/api/insights", methods=["GET"])
def api_insights():
    domain = (request.args.get("domain") or "").upper()
    app_name = request.args.get("app") or ""
    file = request.args.get("file") or ""  # optional: specific summary filename
    controller = request.args.get("controller") or ""

    if domain not in ("APM", "BRUM", "MRUM") or not app_name:
        return jsonify({"error": "Missing domain or app."}), 400

    folder = RESULT_FOLDER

    # Choose file: specific or latest for domain.
    if file:
        path = os.path.join(folder, file)
    else:
        latest_name = latest_summary_file(domain, controller or None)
        path = os.path.join(folder, latest_name) if latest_name else ""

    if not path or not os.path.exists(path):
        return jsonify({"error": "Snapshot not found."}), 404

    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    apps_index = (data.get("appsIndex") or {})
    entry = apps_index.get(app_name)
    if not entry:
        # Try a normalized match
        key = app_name.strip().lower()
        for k, v in apps_index.items():
            if k.strip().lower() == key:
                entry = v
                break

    if not entry:
        return jsonify({"error": "App not found in snapshot."}), 404

    areas = entry.get("areas", [])
    detail = entry.get("detail", {})

    return jsonify(
        {
            "domain": domain,
            "app": app_name,
            "areas": areas,
            "detail": detail,
            "meta": data.get("meta", {}),
        }
    )



@app.route("/api/trends/runs", methods=["GET"])
def api_trends_runs():
    domain = (request.args.get("domain") or "").lower()
    if domain not in ("apm", "brum", "mrum"):
        return jsonify({"error": "Invalid domain."}), 400

    controller = request.args.get("controller")
    try:
        limit = int(request.args.get("limit", "20"))
    except ValueError:
        limit = 20

    baseline = (request.args.get("baseline") or "").lower()

    # 🔹 Use the module-level RESULT_FOLDER (same as other APIs)
    folder = RESULT_FOLDER

    runs = scan_runs(folder, domain=domain, controller_filter=controller, limit=1000)

    if baseline == "earliestprev":
        runs = fixed_baseline_runs(runs)

    runs = runs[:limit]

    series = [
        {
            "compareDate": r["compareDate"],
            "previousDate": r["previousDate"],
            "currentDate": r["currentDate"],
            "improved": r["improved"],
            "degraded": r["degraded"],
            "percentage": r["percentage"],
            "tiers": r["tiers"],
            "file": r["file"],
        }
        for r in runs
    ]

    label = controller or (runs[0]["controller"] if runs else None)
    return jsonify(
        {
            "domain": domain.upper(),
            "controller": label,
            "count": len(series),
        "trendMode": "fixedBaseline" if baseline == "earliestprev" else "latestRuns",
        "baselineDate": runs[0]["previousDate"] if baseline == "earliestprev" and runs else None,
        "items": series,
        }
    )


@app.route("/api/trends/app", methods=["GET"])
def api_trends_app():
    domain = (request.args.get("domain") or "").lower()
    app_name = request.args.get("app") or ""
    controller = request.args.get("controller")

    if domain not in ("apm", "brum", "mrum"):
        return jsonify({"error": "Invalid domain."}), 400
    if not app_name:
        return jsonify({"error": "Missing app."}), 400

    folder = RESULT_FOLDER
    prefix = f"analysis_summary_{domain}_"
    items = []

    if not os.path.isdir(folder):
        return jsonify({"domain": domain.upper(), "controller": controller, "app": app_name, "items": []})

    app_key = app_name.strip().lower()

    for name in os.listdir(folder):
        if not (name.startswith(prefix) and name.endswith(".json")):
            continue

        path = os.path.join(folder, name)
        try:
            with open(path, "r", encoding="utf-8") as f:
                payload = json.load(f)
        except Exception:
            continue

        meta = payload.get("meta") or {}
        if controller and _slug(meta.get("controller")) != _slug(controller):
            continue

        apps_index = payload.get("appsIndex") or {}
        entry = apps_index.get(app_name)
        if not entry:
            for key, value in apps_index.items():
                if key.strip().lower() == app_key:
                    entry = value
                    break
        if not entry:
            continue

        upgraded = 0
        downgraded = 0
        unchanged = 0
        for area in entry.get("areas") or []:
            status = str(area.get("status") or "").lower()
            if status == "upgraded":
                upgraded += 1
            elif status == "downgraded":
                downgraded += 1
            else:
                unchanged += 1

        items.append(
            {
                "file": name,
                "controller": meta.get("controller"),
                "previousDate": meta.get("previousDate") or "",
                "currentDate": meta.get("currentDate") or "",
                "compareDate": meta.get("compareDate") or "",
                "upgraded": upgraded,
                "downgraded": downgraded,
                "unchanged": unchanged,
            }
        )

    items = fixed_baseline_runs(items)
    return jsonify(
        {
            "domain": domain.upper(),
            "controller": controller or (items[0]["controller"] if items else None),
            "app": app_name,
            "trendMode": "fixedBaseline",
            "baselineDate": items[0]["previousDate"] if items else None,
            "count": len(items),
            "items": items,
        }
    )

if __name__ == "__main__":
    def open_browser():
        webbrowser.open("http://127.0.0.1:5000")

    # Give Flask a moment to start, then open the browser
    threading.Timer(1.0, open_browser).start()
    app.run(host="127.0.0.1", port=5000, debug=False)
